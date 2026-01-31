import os
import json
import re
from typing import List, Dict, Any, Tuple

from google.cloud import documentai_v1 as documentai
from google import genai

from rapidfuzz import fuzz
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas


# ================== DIRECTORIES ==================
INPUT_DIR = "input"
OUTPUT_DIR = "output"

os.makedirs(INPUT_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)


# ================== CONFIG ==================
DOC_PROJECT = "636776350532"
DOC_LOCATION = "us"
DOC_PROCESSOR = "ac3490fa842041e5"

GENAI_PROJECT = "finmate-svc-1-88322725"
GENAI_LOCATION = "us-central1"
GENAI_MODEL = "gemini-2.5-flash"
# =================================================


# ================== HELPERS ==================
def die(msg):
    raise RuntimeError(msg)


def assert_all_strings(name: str, seq):
    for i, x in enumerate(seq):
        if not isinstance(x, str):
            die(f"{name}[{i}] IS {type(x)} → {x}")


def find_input_files():
    pdf = None
    excel = None
    for f in os.listdir(INPUT_DIR):
        lf = f.lower()
        if lf.endswith(".pdf"):
            pdf = os.path.join(INPUT_DIR, f)
        elif lf.endswith(".xlsx"):
            excel = os.path.join(INPUT_DIR, f)

    if not pdf or not excel:
        die("input/ must contain one PDF and one XLSX")

    return pdf, excel


# ================== DOCUMENT AI ==================
docai = documentai.DocumentProcessorServiceClient()


def safe_text(doc, anchor) -> str:
    if not anchor or not anchor.text_segments:
        return ""
    out = []
    for seg in anchor.text_segments:
        if seg.start_index is not None and seg.end_index is not None:
            out.append(doc.text[seg.start_index:seg.end_index])
    return "".join(out).strip()


def parse_document(path: str) -> List[str]:
    print("\n[1] OCR started")

    with open(path, "rb") as f:
        content = f.read()

    name = docai.processor_path(DOC_PROJECT, DOC_LOCATION, DOC_PROCESSOR)

    result = docai.process_document(
        request=documentai.ProcessRequest(
            name=name,
            raw_document=documentai.RawDocument(
                content=content,
                mime_type="application/pdf"
            )
        )
    )

    doc = result.document
    values = []

    for page in doc.pages:
        for p in page.paragraphs:
            t = safe_text(doc, p.layout.text_anchor)
            if t:
                values.append(t)

        for table in page.tables:
            for row in table.body_rows:
                for cell in row.cells:
                    t = safe_text(doc, cell.layout.text_anchor)
                    if t:
                        values.append(t)

    assert_all_strings("ocr_values", values)
    print(f"[1] OCR extracted {len(values)} values")
    return values


# ================== EXCEL ==================
def extract_labels_with_positions(path: str) -> Dict[str, Tuple[int, int]]:
    wb = load_workbook(path)
    ws = wb.active

    labels = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                label = str(cell.value).strip()
                if 0 < len(label) < 80:
                    labels[label] = (cell.row, cell.column)

    assert_all_strings("labels", labels.keys())
    return labels


def safe_write(ws, row: int, col: int, value):
    """
    Correct merged-cell-safe writer.
    """
    cell = ws.cell(row=row, column=col)

    # Normal cell
    if not isinstance(cell, MergedCell):
        cell.value = value
        return

    # Merged cell → write to top-left owner
    for rng in ws.merged_cells.ranges:
        if (
            rng.min_row <= row <= rng.max_row
            and rng.min_col <= col <= rng.max_col
        ):
            ws.cell(
                row=rng.min_row,
                column=rng.min_col
            ).value = value
            return

    # Fallback
    ws.cell(row=row, column=col).value = value


# ================== MATCHING ==================
def generate_candidates(label: str, ocr_values: List[str], top_k=5):
    scored = []
    for v in ocr_values:
        score = fuzz.partial_ratio(label.lower(), v.lower())
        if score >= 40:
            scored.append((score, v))
    scored.sort(reverse=True)
    return [v for _, v in scored[:top_k]]


genai_client = genai.Client(
    vertexai=True,
    project=GENAI_PROJECT,
    location=GENAI_LOCATION,
)


def gemini_choose(label: str, candidates: List[str]) -> Dict[str, Any]:
    if not candidates:
        return {"value": None, "confidence": 0.0}

    prompt = f"""
Pick the correct value.

Field:
{label}

Candidates:
{candidates}

Return ONLY JSON:
{{"value":"...","confidence":0.0}}
"""

    resp = genai_client.models.generate_content(
        model=GENAI_MODEL,
        contents=prompt,
    )

    match = re.search(r"\{[\s\S]*?\}", str(resp.text))
    parsed = json.loads(match.group())

    if parsed["value"] not in candidates:
        parsed["value"] = candidates[0]
        parsed["confidence"] = 0.5

    return parsed


# ================== PDF ==================
def write_pdf(results: Dict[str, Any], path: str):
    c = canvas.Canvas(path, pagesize=A4)
    w, h = A4
    y = h - 40

    c.setFont("Helvetica-Bold", 14)
    c.drawString(40, y, "Extraction Summary")
    y -= 30

    c.setFont("Helvetica", 10)
    for k, v in results.items():
        if y < 80:
            c.showPage()
            y = h - 40
        c.drawString(40, y, f"{k}: {v['value']} ({v['confidence']})")
        y -= 18

    c.save()


# ================== MAIN ==================
def main():
    print("\n=== TERMINAL PIPELINE STARTED ===")

    pdf, excel = find_input_files()
    ocr_values = parse_document(pdf)

    labels = extract_labels_with_positions(excel)
    wb = load_workbook(excel)
    ws = wb.active

    results = {}

    for label, (r, c) in labels.items():
        candidates = generate_candidates(label, ocr_values)
        chosen = gemini_choose(label, candidates)
        safe_write(ws, r, c + 1, chosen["value"])
        results[label] = chosen

    excel_out = os.path.join(OUTPUT_DIR, "filled.xlsx")
    pdf_out = os.path.join(OUTPUT_DIR, "summary.pdf")

    wb.save(excel_out)
    write_pdf(results, pdf_out)

    print("\n✅ OUTPUT GENERATED SUCCESSFULLY")
    print("Excel:", excel_out)
    print("PDF:", pdf_out)


if __name__ == "__main__":
    main()
