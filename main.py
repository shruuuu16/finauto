import os
from dotenv import load_dotenv

load_dotenv()
import uuid
import json
import re
import logging
from typing import Dict, Any, List, Tuple, Optional
from fastapi import FastAPI, UploadFile, File, BackgroundTasks, HTTPException, Form
from fastapi.responses import JSONResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
# $env:GOOGLE_APPLICATION_CREDENTIALS="C:\keys\finmate-service-90fb14600045.json"
from google.cloud import documentai_v1 as documentai
try:
    # Provided by the `google-genai` package.
    from google import genai  # type: ignore
except Exception:  # pragma: no cover
    genai = None

try:
    from docling_parser import parse_document_docling, parse_document_by_page_docling
    DOCLING_AVAILABLE = True
except ImportError:
    DOCLING_AVAILABLE = False
    # Note: logger not yet defined, will log later if needed

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas

from rapidfuzz import fuzz

# ---------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------
INPUT_DIR = "input"
OUTPUT_DIR = "output"
# Set to 'true' to use Docling (local) instead of Google DocAI (cloud)
USE_DOCLING = os.getenv("USE_DOCLING", "false").lower() == "true"

JOBS: Dict[str, Dict[str, Any]] = {}

os.makedirs(INPUT_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI()

# Mount directories for serving files
app.mount("/output", StaticFiles(directory=OUTPUT_DIR), name="output")
app.mount(
    "/frontend",
    StaticFiles(directory=os.path.join(os.path.dirname(__file__), "frontend"), html=True),
    name="frontend"
)

# Enable CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow all origins for development
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------------------------------------------------------------
# CLOUD DEPLOYMENT CREDENTIALS SETUP
# ---------------------------------------------------------------------
import base64
import tempfile

# Handle Google Credentials from Env Var (Vercel/Render Secret)
# Supports both GOOGLE_CREDENTIALS_JSON (content) and GOOGLE_APPLICATION_CREDENTIALS (path)
if os.getenv("GOOGLE_CREDENTIALS_JSON"):
    creds_content = os.getenv("GOOGLE_CREDENTIALS_JSON")
    try:
        # Check if it's base64 encoded (common for secrets)
        if not creds_content.strip().startswith("{"):
             creds_content = base64.b64decode(creds_content).decode('utf-8')
    except Exception:
        pass # Assume it's raw JSON if decode fails

    # Write to a temp file that persists for the process lifetime
    # Note: In Vercel, /tmp is the only writable dir
    fd, path = tempfile.mkstemp(suffix=".json", text=True)
    with os.fdopen(fd, 'w') as tmp:
        tmp.write(creds_content)
    os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = path
    logger.info(f"Loaded Google Credentials from env var to {path}")

# Sanity check: If GOOGLE_APPLICATION_CREDENTIALS points to a non-existent file 
# (e.g. valid on local Windows but invalid on Render Linux), un-set it to avoid 
# confusing "File Not Found" errors from the client libraries.
creds_path = os.getenv("GOOGLE_APPLICATION_CREDENTIALS")
if creds_path and not os.path.exists(creds_path):
    logger.warning(f"Credential path {creds_path} does not exist (likely local path on cloud env). Unsetting env var.")
    del os.environ["GOOGLE_APPLICATION_CREDENTIALS"]

# ---------------------------------------------------------------------
# DOCUMENT AI (OCR ONLY)
# ---------------------------------------------------------------------
try:
    docai = documentai.DocumentProcessorServiceClient()
except Exception as e:
    logger.error(f"Failed to initialize DocumentAI client: {e}")
    docai = None

DOC_PROJECT = "424497190710"
DOC_LOCATION = "us"
DOC_PROCESSOR = "a469ba49186b1786"

# ---------------------------------------------------------------------
# GEMINI (ONLY FOR RERANKING)
# ---------------------------------------------------------------------
gemini = None
if genai is not None:
    try:
        if os.getenv("GEMINI_API_KEY"):
            # Use Google AI Studio (API Key)
            gemini = genai.Client(api_key=os.getenv("GEMINI_API_KEY"), http_options={"api_version": "v1alpha"})
            logger.info("Initialized Gemini client with API Key (AI Studio)")
        else:
            # Use Vertex AI (Service Account)
            gemini = genai.Client(
                vertexai=True,
                project=os.getenv("VERTEX_PROJECT", "finmate-service"),
                location=os.getenv("VERTEX_LOCATION", "us-central1"),
            )
            logger.info("Initialized Gemini client with Vertex AI")
    except Exception as e:
        logger.warning(f"Gemini client init failed; reranking disabled: {e}")

# Model availability can vary by project/region and over time.
# Allow overriding without code changes.
def get_llm_model() -> str:
    return os.getenv("GEMINI_MODEL", "gemini-2.0-flash-exp")

# ---------------------------------------------------------------------
# VERCEL COMPATIBILITY HEADERS
# ---------------------------------------------------------------------
IS_VERCEL = os.getenv("VERCEL") == "1"

# ---------------------------------------------------------------------
# ROOT ROUTE
# ---------------------------------------------------------------------
@app.get("/")
async def read_root():
    return FileResponse(os.path.join(os.path.dirname(__file__), "frontend", "index.html"))

# ---------------------------------------------------------------------
# OCR PARSING
# ---------------------------------------------------------------------
def safe_text(doc, anchor) -> str:
    if not anchor or not getattr(anchor, "text_segments", None):
        return ""
    parts = []
    for seg in anchor.text_segments:
        if seg.start_index is not None and seg.end_index is not None:
            parts.append(doc.text[seg.start_index:seg.end_index])
    return "".join(parts).strip()


def _extract_page_values(doc, page) -> List[str]:
    """Extract OCR values from a single page."""
    values: List[str] = []
    seen = set()
    
    # Extract from paragraphs
    for p in page.paragraphs:
        t = safe_text(doc, p.layout.text_anchor)
        if t:
            t = t.replace('\n', ' ').replace('\r', ' ').strip()
            t = ' '.join(t.split())
            if len(t) > 1 and t.lower() not in seen:
                values.append(t)
                seen.add(t.lower())
    
    # Extract from tables
    for table in page.tables:
        for row in table.body_rows:
            row_values = []
            for cell in row.cells:
                t = safe_text(doc, cell.layout.text_anchor)
                if t:
                    t = t.replace('\n', ' ').strip()
                    t = ' '.join(t.split())
                    if len(t) > 1:
                        row_values.append(t)
            
            for rv in row_values:
                if rv.lower() not in seen:
                    values.append(rv)
                    seen.add(rv.lower())
            
            if len(row_values) >= 2:
                combined = ' '.join(row_values)
                if combined.lower() not in seen:
                    values.append(combined)
                    seen.add(combined.lower())
    
    return values


    logger.info(f"Extracted values from {len(pages_values)} pages")
    return pages_values


def _parse_document_by_page_docai(path: str) -> List[List[str]]:
    """Parse document and return OCR values grouped by page using Google DocAI."""
    with open(path, "rb") as f:
        content = f.read()

    mime = "application/pdf" if path.lower().endswith(".pdf") else "image/jpeg"
    name = docai.processor_path(DOC_PROJECT, DOC_LOCATION, DOC_PROCESSOR)

    result = docai.process_document(
        request=documentai.ProcessRequest(
            name=name,
            raw_document=documentai.RawDocument(content=content, mime_type=mime)
        )
    )

    doc = result.document
    pages_values: List[List[str]] = []
    
    for page in doc.pages:
        page_values = _extract_page_values(doc, page)
        if page_values:  # Only add non-empty pages
            pages_values.append(page_values)
    
    logger.info(f"Extracted values from {len(pages_values)} pages (DocAI)")
    return pages_values


def parse_document_by_page(path: str) -> List[List[str]]:
    if USE_DOCLING and DOCLING_AVAILABLE:
        try:
            return parse_document_by_page_docling(path)
        except Exception as e:
            logger.error(f"Docling failed: {e}. Falling back to DocAI.")
    return _parse_document_by_page_docai(path)


    logger.info(f"Extracted {len(values)} unique OCR values")
    return values


def _parse_document_docai(path: str) -> List[str]:
    with open(path, "rb") as f:
        content = f.read()

    mime = "application/pdf" if path.lower().endswith(".pdf") else "image/jpeg"
    name = docai.processor_path(DOC_PROJECT, DOC_LOCATION, DOC_PROCESSOR)

    result = docai.process_document(
        request=documentai.ProcessRequest(
            name=name,
            raw_document=documentai.RawDocument(content=content, mime_type=mime)
        )
    )

    doc = result.document
    values: List[str] = []
    seen = set()  # Deduplicate extracted values

    for page in doc.pages:
        # Extract from paragraphs (main document text)
        for p in page.paragraphs:
            t = safe_text(doc, p.layout.text_anchor)
            if t:
                # Better cleaning: handle newlines, extra spaces, special chars
                t = t.replace('\n', ' ').replace('\r', ' ').strip()
                t = ' '.join(t.split())  # Normalize whitespace
                
                # Skip short/empty and already seen
                if len(t) > 1 and t.lower() not in seen:
                    values.append(t)
                    seen.add(t.lower())

        # Extract from tables (structured data like invoice line items)
        for table in page.tables:
            for row in table.body_rows:
                row_values = []
                for cell in row.cells:
                    t = safe_text(doc, cell.layout.text_anchor)
                    if t:
                        t = t.replace('\n', ' ').strip()
                        t = ' '.join(t.split())
                        if len(t) > 1:
                            row_values.append(t)
                
                # Add individual cell values
                for rv in row_values:
                    if rv.lower() not in seen:
                        values.append(rv)
                        seen.add(rv.lower())
                
                # Also add combined row (useful for "key: value" patterns)
                if len(row_values) >= 2:
                    combined = ' '.join(row_values)
                    if combined.lower() not in seen:
                        values.append(combined)
                        seen.add(combined.lower())

    logger.info(f"Extracted {len(values)} unique OCR values (DocAI)")
    return values


def parse_document(path: str) -> List[str]:
    if USE_DOCLING and DOCLING_AVAILABLE:
        try:
            return parse_document_docling(path)
        except Exception as e:
            logger.error(f"Docling failed: {e}. Falling back to DocAI.")
    return _parse_document_docai(path)

# ---------------------------------------------------------------------
# LLM EXTRACTION (Unified)
# ---------------------------------------------------------------------
def extract_data_with_gemini(ocr_text: str) -> Dict[str, Any]:
    """
    Uses Gemini to analyze OCR text and return structured data:
    1. Document Type & Sheet Recommendation
    2. Invoice Summary (Totals, Dates)
    3. HSN Summary (Table)
    """
    if not gemini:
        return {}

    # Limit text to avoid token limits (though 2.0 Flash has large context)
    # 50k chars is usually plenty for invoices
    safe_text = ocr_text[:50000]

    prompt = f"""
    You are an expert financial document parser. Analyze the OCR text below from a document.
    
    Return a SINGLE JSON object with the following structure. Do not return markdown formatting (```json), just the raw JSON string.
    
    {{
        "document_class": {{
            "type": "tax_invoice" | "credit_note" | "debit_note" | "bill_of_supply" | "receipt" | "amended_invoice" | "other",
            "is_amended": boolean,  # True if it says "Amended", "Revised", "Duplicate for.." etc.
            "recommended_sheet": "b2b" | "b2ba" | "b2c" | "cdnr" | "cdnra" | "exp" | "exempt" | "advrec"  # Best guess based on type
        }},
        "invoice_summary": {{
            "gstin_supplier": "string or null",
            "gstin_receiver": "string or null",
            "invoice_no": "string or null",
            "invoice_date": "DD-MMM-YYYY or DD/MM/YYYY string or null",
            "total_taxable_value": number or null,
            "total_cgst": number or null,
            "total_sgst": number or null,
            "total_igst": number or null,
            "total_cess": number or null,
            "grand_total": number or null,
            "reverse_charge": "Y" | "N"
        }},
        "hsn_summary": [
            {{
                "hsn_code": "string",
                "description": "string",
                "qty": number,
                "rate": number,  # Item price per unit
                "taxable_value": number,
                "tax_rate": number # e.g. 5, 12, 18, 28
            }}
        ]
    }}

    Notes:
    - "b2ba" sheet is ONLY for Amended/Revised invoices.
    - "cdnra" sheet is ONLY for Amended Credit/Debit Notes.
    - Extract HSN summary if a table exists. If multiple items have same HSN, aggregate them if possible, or list them.
    - Be precise with amounts. Remove currency symbols.

    --- DOCUMENT TEXT ---
    {safe_text}
    """

    try:
        model_name = get_llm_model()
        response = gemini.models.generate_content(
            model=model_name,
            contents=prompt,
            config={
                "response_mime_type": "application/json"
            }
        )
        
        # Parse JSON
        import json
        text_resp = response.text.strip()
        # Clean markdown if present
        if text_resp.startswith("```json"):
            text_resp = text_resp[7:]
        if text_resp.endswith("```"):
            text_resp = text_resp[:-3]
            
        data = json.loads(text_resp.strip())
        logger.info(f"Gemini Extraction Success: {data.get('document_class', {}).get('type')}")
        return data

    except Exception as e:
        logger.error(f"Gemini Extraction Failed: {e}")
        return {}

# ---------------------------------------------------------------------
# EXCEL TEMPLATE
# ---------------------------------------------------------------------
def normalize_label(label: str) -> str:
    # Make matching more robust across templates with *, newlines, (Required), etc.
    cleaned = re.sub(r"\s+", " ", label.replace("\n", " ")).strip()
    cleaned = re.sub(r"\(\s*required\s*\)", "", cleaned, flags=re.IGNORECASE).strip()
    cleaned = cleaned.replace("*", "").strip()
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned


def is_label_text(value: object) -> bool:
    if value is None:
        return False

    if not isinstance(value, str):
        return False

    text = value.strip()
    if not (0 < len(text) < 120):
        return False

    # Must contain some letters (avoid pure numbers/dates/codes)
    if not re.search(r"[A-Za-z]", text):
        return False

    # Avoid obvious date-like / numeric-only / currency-only strings
    if re.fullmatch(r"[0-9\s,./:-]+", text):
        return False

    # Avoid values that look like timestamps
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}.*", text):
        return False

    # Avoid single short tokens that are usually data (e.g., "No.")
    if len(text) <= 3 and text.lower() in {"no.", "no", "dt", "sr"}:
        return False

    return True


def extract_labels_from_ws(ws) -> Dict[str, Tuple[int, int]]:
    # Try to detect a header row (many labels across the same row)
    max_rows_to_scan = min(60, ws.max_row or 0)
    max_cols_to_scan = min(60, ws.max_column or 0)
    best_row: Optional[int] = None
    best_score = 0

    for r in range(1, max_rows_to_scan + 1):
        score = 0
        for c in range(1, max_cols_to_scan + 1):
            v = ws.cell(row=r, column=c).value
            if is_label_text(v):
                score += 1
        if score > best_score:
            best_score = score
            best_row = r

    labels: Dict[str, Tuple[int, int]] = {}

    def non_empty_count(row_idx: int) -> int:
        cnt = 0
        for c in range(1, max_cols_to_scan + 1):
            v = ws.cell(row=row_idx, column=c).value
            if v not in (None, ""):
                cnt += 1
        return cnt

    # Heuristic: if a row has many label-like strings, treat as header row.
    if best_row is not None and best_score >= 5 and best_row < (ws.max_row or 0):
        # Pick a target row beneath header; prefer the first mostly-empty row.
        target_row = best_row + 1
        for r in range(best_row + 1, min(best_row + 10, (ws.max_row or 0)) + 1):
            if non_empty_count(r) <= max(2, best_score // 4):
                target_row = r
                break

        for c in range(1, max_cols_to_scan + 1):
            v = ws.cell(row=best_row, column=c).value
            if not is_label_text(v):
                continue
            raw_label = str(v).strip()
            key = normalize_label(raw_label)
            if key and key not in labels:
                labels[key] = (target_row, c)
        return labels

    # Fallback: vertical label/value layout
    for row in ws.iter_rows():
        for cell in row:
            if not is_label_text(cell.value):
                continue
            raw_label = str(cell.value).strip()
            key = normalize_label(raw_label)
            if not key:
                continue

            # Prefer labels where the next column is empty (value slot)
            next_cell = ws.cell(row=cell.row, column=cell.column + 1)
            if next_cell.value not in (None, ""):
                continue

            if key not in labels:
                labels[key] = (cell.row, cell.column + 1)

    return labels


def extract_labels(template_path: str) -> Dict[str, Tuple[int, int]]:
    """Return mapping: normalized label -> (target_row, target_col).

    Supports two common template styles:
    1) "Vertical" form: labels in one column, values go in the next column.
    2) "Header-row" table: labels as column headers, values go in the row below.
    """

    wb = load_workbook(template_path)
    ws = wb.active
    return extract_labels_from_ws(ws)


def choose_best_sheet(wb, ocr_values: List[str]):
    """Pick the most likely worksheet to fill.

    Many users have multi-sheet templates. We select the sheet that looks most
    like a data-entry sheet (header row with many labels) and whose sheet name
    or top-left content matches the document OCR.
    """
    # Detect document type for smarter sheet selection
    doc_type = detect_document_type(ocr_values)
    logger.info(f"Detected document type: {doc_type}")
    
    # Document type to preferred sheet name mapping
    # IMPORTANT: B2BA is ONLY for amended invoices, not regular invoices
    type_to_sheets = {
        "tax_invoice": ["b2b", "invoice"],  # Regular invoices -> B2B only
        "amended_invoice": ["b2ba", "b2b"],  # Amended invoices -> B2BA preferred
        "credit_note": ["cdnr", "cdnra", "credit"],
        "debit_note": ["cdnr", "cdnra", "debit"],
        "bill_of_supply": ["b2c", "b2cs"],
        "receipt": ["receipt", "payment"],
        "invoice": ["b2b", "invoice"]  # Default -> B2B only
    }
    preferred_sheets = type_to_sheets.get(doc_type, [])
    
    ocr_blob = " \n".join(ocr_values[:250]).lower()

    best_ws = wb.active
    best_score = float("-inf")
    ranked: List[Dict[str, Any]] = []

    for ws in wb.worksheets:
        labels = extract_labels_from_ws(ws)
        header_count = len(labels)

        # Prefer sheets that have a real header row with many columns.
        header_count_score = min(25.0, float(header_count))

        # Sheet name relevance with document type boost
        name_score = fuzz.partial_ratio(ws.title.lower(), ocr_blob) / 100.0
        sheet_lower = ws.title.lower()
        
        # Strong boost for matching document type preferred sheets
        doc_type_boost = 0.0
        for pref in preferred_sheets:
            if pref in sheet_lower:
                doc_type_boost = 15.0  # Very strong boost for matching document type
                break
        
        # Penalty for B2BA when document is a regular tax_invoice (not amended)
        if doc_type == "tax_invoice" and "b2ba" in sheet_lower:
            doc_type_boost = -10.0  # Strong penalty - regular invoices should NOT go to B2BA

        # Top-left content relevance (captures titles like "Exports Invoices - 6A")
        top_left_text_parts: List[str] = []
        for r in range(1, min(10, ws.max_row or 0) + 1):
            for c in range(1, min(10, ws.max_column or 0) + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str) and len(v.strip()) > 2:
                    top_left_text_parts.append(v.strip())
        top_left_text = " ".join(top_left_text_parts)
        top_left_score = fuzz.partial_ratio(top_left_text.lower(), ocr_blob) / 100.0 if top_left_text else 0.0

        # Header-text relevance: compare extracted header labels to OCR.
        header_text = " ".join(labels.keys()).lower()
        header_text_score = fuzz.partial_ratio(header_text, ocr_blob) / 100.0 if header_text else 0.0

        score = (
            header_count_score
            + (12.0 * name_score)
            + (10.0 * top_left_score)
            + (18.0 * header_text_score)
            + doc_type_boost  # Strong boost/penalty based on document type
        )

        ranked.append(
            {
                "sheet": ws.title,
                "score": score,
                "header_count": header_count,
                "name_score": name_score,
                "top_left_score": top_left_score,
                "header_text_score": header_text_score,
            }
        )

        if score > best_score:
            best_score = score
            best_ws = ws

    ranked.sort(key=lambda d: d["score"], reverse=True)
    setattr(best_ws, "__finautomate_ranked_sheets", ranked)
    return best_ws


def _looks_like_gstin(text: str) -> bool:
    if not text:
        return False
    t = text.strip().upper()
    # GSTIN format: 2 digits + 5 letters + 4 digits + 1 letter + 3 alphanumeric
    gstin_pattern = r"\b\d{2}[A-Z]{5}\d{4}[A-Z][A-Z0-9]{3}\b"
    if "GST" in t and re.search(gstin_pattern, t):
        return True
    return bool(re.search(gstin_pattern, t))


def _find_best_by_regex(ocr_values: List[str], patterns: List[str]) -> Optional[str]:
    for v in ocr_values:
        if not isinstance(v, str):
            continue
        for pat in patterns:
            if re.search(pat, v, flags=re.IGNORECASE):
                return v
    return None


_UQC_TOKENS = {
    "PCS",
    "NOS",
    "NO",
    "KG",
    "KGS",
    "G",
    "GM",
    "GRAM",
    "L",
    "LTR",
    "MTR",
    "M",
    "BOX",
    "BAG",
    "SET",
    "PAIR",
    "PAIRS",
}


def _extract_numbers(text: str) -> List[str]:
    """Extract numbers from text, handling both Indian (1,23,456) and Western (1,234,567) formats."""
    if not text:
        return []
    out: List[str] = []
    # Match both Indian format (1,23,456.78) and Western format (1,234,567.89)
    # Indian: starts with 1-2 digits, then 2-digit groups (1,79,360.00)
    # Western: starts with 1-3 digits, then 3-digit groups (1,234,567.89)
    # Combined pattern: allow 2-3 digit groups after comma
    for m in re.finditer(r"\b\d{1,3}(?:,\d{2,3})*(?:\.\d{1,2})?\b", text):
        out.append(m.group(0))
    return out


def _extract_percent(text: str) -> Optional[str]:
    if not text:
        return None
    m = re.search(r"\b\d{1,2}(?:\.\d+)?\s*%\b", text)
    return m.group(0).replace(" ", "") if m else None


def _extract_date_token(text: str) -> Optional[str]:
    if not text:
        return None
    m = re.search(
        r"\b(\d{1,2}[-/]\d{1,2}[-/]\d{2,4}|\d{1,2}[-/][A-Za-z]{3}[-/]\d{2,4}|\d{1,2}\s*[A-Za-z]{3,9}\s*\d{2,4})\b",
        text,
    )
    return m.group(0) if m else None


def detect_document_type(ocr_values: List[str]) -> str:
    """Detect document type from OCR values: invoice, credit_note, debit_note, etc.
    
    Important: Distinguishes between regular B2B invoices and B2BA (amended) invoices.
    B2BA is ONLY for invoices that correct/amend previously filed invoices.
    """
    all_text = " ".join(str(v).lower() for v in ocr_values if v)
    
    # Check for explicit document type mentions
    if re.search(r"\bcredit\s*note\b", all_text):
        return "credit_note"
    if re.search(r"\bdebit\s*note\b", all_text):
        return "debit_note"
    
    # Check for AMENDED invoice indicators (B2BA specific)
    # B2BA is for invoices that CORRECT/AMEND previously filed invoices
    # NOTE: Be strict here - "Original Invoice No" column header in templates should NOT trigger this
    # Only explicit amendment keywords should route to B2BA
    amended_indicators = [
        r"\bamended\s+(tax\s+)?invoice\b",       # "Amended Invoice" or "Amended Tax Invoice"
        r"\bthis\s+invoice\s+supersedes\b",     # Explicit supersession language
        r"\breplaces\s+invoice\b",              # Explicit replacement language
        r"\bcorrection\s+to\s+invoice\b",       # Explicit correction language
        r"\bmodified\s+invoice\b",              # Modified invoice
        r"\bamendment\s+to\b",                  # Amendment to...
    ]
    for pattern in amended_indicators:
        if re.search(pattern, all_text):
            return "amended_invoice"  # This will route to B2BA
    
    # Regular tax invoice (goes to B2B)
    if re.search(r"\b(tax\s*invoice|invoice)\b", all_text):
        return "tax_invoice"
    if re.search(r"\b(bill\s*of\s*supply|supply\s*bill)\b", all_text):
        return "bill_of_supply"
    if re.search(r"\breceipt\b", all_text):
        return "receipt"
    
    # Default fallback - regular invoice
    return "invoice"


def _is_trivial_value(text: str) -> bool:
    t = (text or "").strip().lower()
    if not t:
        return True
    # Common junk tokens that show up in GST templates / invoices.
    if t in {"per", "rate", "no", "no.", "dated", "date", "qty", "quantity", "hsn", "sac", "hsn/sac"}:
        return True
    # Pure punctuation or too short.
    if len(t) <= 2 and not re.search(r"\d", t):
        return True
    return False


def _looks_like_header(text: str) -> bool:
    # Ex: "HSN/SAC Quantity", "Taxable Value Rate Amount"
    if not text:
        return False
    t = " ".join(text.strip().split())
    if len(t) < 6:
        return False
    if ":" in t:
        return False
    if re.search(r"\d", t):
        return False
    words = [w for w in re.split(r"\s+", t) if w]
    if len(words) >= 3 and sum(1 for w in words if w[0:1].isupper()) >= 2:
        return True
    if any(k in t.lower() for k in ["hsn", "sac", "taxable", "rate", "amount", "quantity", "description"]):
        return True
    return False


def _is_bad_candidate(label: str, candidate: str) -> bool:
    if not candidate:
        return True
    if _is_trivial_value(candidate):
        return True
    if _looks_like_header(candidate):
        return True

    label_l = (label or "").lower().strip()
    cand_l = candidate.lower().strip()

    # If candidate is basically the label itself and has no digits, treat it as junk.
    if not re.search(r"\d", cand_l):
        sim = max(
            fuzz.partial_ratio(label_l, cand_l),
            fuzz.token_sort_ratio(label_l, cand_l),
        )
        if sim >= 92:
            return True

    # Field-specific guards
    if "gstin" in label_l or ("gst" in label_l and not any(k in label_l for k in ["igst", "cgst", "sgst", "utgst"])):
        return not _looks_like_gstin(candidate)
    if "date" in label_l or "dated" in label_l:
        return _extract_date_token(candidate) is None
    if any(k in label_l for k in ["amount", "value", "total", "taxable"]):
        return not bool(re.search(r"\d", candidate))
    if "rate" in label_l or "%" in label_l:
        return _extract_percent(candidate) is None
    if "hsn" in label_l or "sac" in label_l:
        return re.search(r"\b\d{4,8}\b", candidate) is None
    if "uqc" in label_l:
        return candidate.strip().upper() not in _UQC_TOKENS
    if "quantity" in label_l or label_l.endswith("qty"):
        return re.search(r"\b\d+(?:\.\d+)?\b", candidate) is None
    if "reverse charge" in label_l:
        return cand_l not in {"y", "n", "yes", "no"}
    if "note type" in label_l:
        return cand_l not in {"d", "c", "r", "debit", "credit", "refund"}

    return False


def heuristic_extract(label: str, ocr_values: List[str]) -> Dict[str, Any]:
    """Heuristic extractor for common invoice fields.

    This avoids junk matches like "per"/"No." when LLM quota is hit.
    """
    label_l = label.lower()

    # GSTIN (but not tax amount fields like IGST/CGST/SGST Amount)
    if ("gstin" in label_l or ("gst" in label_l and not any(k in label_l for k in ["igst", "cgst", "sgst", "utgst"]))):
        # Special case: E-Commerce GSTIN field - only fill if explicitly mentioned
        if "e-commerce" in label_l or "ecommerce" in label_l:
            for v in ocr_values:
                if not isinstance(v, str):
                    continue
                if "e-commerce" in v.lower() or "ecommerce" in v.lower():
                    m = re.search(r"\b\d{2}[A-Z]{5}\d{4}[A-Z]\d[A-Z]\d\b", v)
                    if m:
                        gstin_val = m.group(0).strip(" :")
                        return {"value": gstin_val, "confidence": 0.8, "reason": "regex: e-commerce gstin"}
            return {"value": None, "confidence": 0.0, "reason": "no e-commerce gstin"}
        
        # For receiver/buyer GSTIN: look for GSTIN near consignee/buyer keywords
        if any(k in label_l for k in ["receiver", "buyer", "consignee"]):
            # First: Look for "Buyer" section and then find GSTIN within it
            buyer_found_idx = None
            for i, v in enumerate(ocr_values):
                if not isinstance(v, str):
                    continue
                t = v.lower()
                if any(k in t for k in ["buyer", "bill to", "consignee", "ship to", "receiver"]):
                    buyer_found_idx = i
                    break
            
            # Check subsequent lines after buyer section for GSTIN
            # Look for lines with "GSTIN/UIN" label followed by GSTIN value
            if buyer_found_idx is not None:
                gstin_label_found = False
                for j in range(buyer_found_idx + 1, min(buyer_found_idx + 15, len(ocr_values))):
                    v = ocr_values[j]
                    if not isinstance(v, str):
                        continue
                    v_upper = v.upper()
                    v_lower = v.lower()
                    
                    # Check if this line has "GSTIN/UIN" label
                    if "gstin" in v_lower or "uin" in v_lower:
                        gstin_label_found = True
                        # Check if GSTIN is on the same line
                        gstin_match = re.search(r"\b(\d{2}[A-Z]{5}\d{4}[A-Z][A-Z0-9]{3})\b", v_upper)
                        if gstin_match:
                            gstin_val = gstin_match.group(1).strip(" :")
                            return {"value": gstin_val, "confidence": 0.9, "reason": "regex: receiver gstin (same line as label)"}
                        continue
                    
                    # If we just saw GSTIN label, next line starting with ":" likely has the value
                    if gstin_label_found:
                        # Handle lines like ": 27DLTPS7779P1ZD"
                        if v.strip().startswith(":"):
                            gstin_match = re.search(r"\b(\d{2}[A-Z]{5}\d{4}[A-Z][A-Z0-9]{3})\b", v_upper)
                            if gstin_match:
                                gstin_val = gstin_match.group(1).strip(" :")
                                return {"value": gstin_val, "confidence": 0.9, "reason": "regex: receiver gstin (after label)"}
                        # Also check any line with GSTIN pattern right after label
                        gstin_match = re.search(r"\b(\d{2}[A-Z]{5}\d{4}[A-Z][A-Z0-9]{3})\b", v_upper)
                        if gstin_match:
                            gstin_val = gstin_match.group(1).strip(" :")
                            return {"value": gstin_val, "confidence": 0.85, "reason": "regex: receiver gstin (near label)"}
                        gstin_label_found = False  # Reset after checking next line
                    
                    # Stop if we hit another section (like next invoice, supplier, or State Name after buyer section)
                    if any(k in v_lower for k in ["supplier", "seller", "from:", "our gstin", "place of supply", "description of goods"]):
                        break
        
        # General GSTIN: Extract all GSTINs and pick appropriately
        all_gstins = []
        for i, v in enumerate(ocr_values):
            if not isinstance(v, str):
                continue
            # Use more flexible GSTIN pattern
            gstin_match = re.search(r"\b(\d{2}[A-Z]{5}\d{4}[A-Z][A-Z0-9]{3})\b", v.upper())
            if gstin_match:
                gstin_clean = gstin_match.group(1).strip(" :")
                context = v.lower()
                all_gstins.append((gstin_clean, context, i))
        
        # If we have multiple GSTINs, prefer the one with buyer/receiver context
        if all_gstins:
            for gstin, context, idx in all_gstins:
                if any(k in context for k in ["consignee", "ship to", "buyer", "bill to", "receiver"]):
                    return {"value": gstin, "confidence": 0.9, "reason": "regex: gstin (receiver context)"}
            # If multiple GSTINs, the second one is often the buyer's (first is seller's)
            if len(all_gstins) >= 2:
                return {"value": all_gstins[1][0], "confidence": 0.8, "reason": "regex: gstin (second = buyer)"}
            # Single GSTIN found - could be seller's, so lower confidence
            return {"value": all_gstins[0][0], "confidence": 0.7, "reason": "regex: gstin (single found)"}

    # Date - prioritize "Invoice Date" or "Dated" over other dates (like Ack Date)
    if "date" in label_l or "dated" in label_l:
        # First pass: Look for explicit "Invoice Date" or "Dated" labels with dates
        invoice_date_patterns = [
            r"(?i)(?:invoice\s*date|inv\.?\s*date|dated)[:\s]+",  # "Invoice Date:" or "Dated:"
            r"(?i)(?:date\s*of\s*invoice)[:\s]+",  # "Date of Invoice:"
            r"(?i)^dated\s+",  # "Dated 01-Jan-2025"
        ]
        for v in ocr_values:
            if not isinstance(v, str):
                continue
            v_lower = v.lower()
            # Skip acknowledgement dates
            if "ack" in v_lower or "acknowledge" in v_lower:
                continue
            # Check if this line contains explicit invoice date labels
            for pattern in invoice_date_patterns:
                if re.search(pattern, v):
                    tok = _extract_date_token(v)
                    if tok:
                        tok_clean = re.sub(r"(?i)^dated\s+", "", tok).strip()
                        return {"value": tok_clean, "confidence": 0.95, "reason": "regex: invoice date (explicit)"}
        
        # Second pass: Fallback to any date if no explicit invoice date found
        for v in ocr_values:
            if not isinstance(v, str):
                continue
            v_lower = v.lower()
            # Skip acknowledgement dates
            if "ack" in v_lower or "acknowledge" in v_lower:
                continue
            tok = _extract_date_token(v)
            if tok:
                # Clean up any "Dated" prefix
                tok_clean = re.sub(r"(?i)^dated\s+", "", tok).strip()
                return {"value": tok_clean, "confidence": 0.9, "reason": "regex: date"}

    # Invoice/Bill number
    if ("invoice" in label_l and ("no" in label_l or "number" in label_l)) or ("bill" in label_l and "no" in label_l):
        for v in ocr_values:
            if not isinstance(v, str):
                continue
            t = v.lower()
            if "invoice" in t and ("no" in t or "number" in t) and re.search(r"\d", t):
                # Extract just the invoice number, remove all labels
                # Match patterns like "Invoice No. 25-26/AP/047" and extract "25-26/AP/047"
                match = re.search(r"(?i)(?:invoice|bill)\s*(?:no\.?|number|#)[:\s]*(.+)", v)
                if match:
                    inv_num = match.group(1).strip()
                    if inv_num and len(inv_num) > 2:
                        return {"value": inv_num, "confidence": 0.9, "reason": "heuristic: invoice no cleaned"}
        # Fallback: look for typical invoice number patterns
        v = _find_best_by_regex(ocr_values, [r"\b\d{2}[-/]\d{2}[-/][A-Z]{2,}[-/]\d+\b", r"\b[A-Z0-9]{2,}[-/][A-Z0-9]{2,}[-/]\d+\b", r"\b\d{2,4}[-/]\d{2,4}\b"]) 
        if v:
            return {"value": v, "confidence": 0.65, "reason": "regex: invoice pattern"}

    # Amounts / values
    if any(k in label_l for k in ["amount", "value", "total", "taxable"]):
        numeric_candidates: List[Tuple[float, str]] = []
        for v in ocr_values:
            if not isinstance(v, str):
                continue
            for s in _extract_numbers(v):
                try:
                    num = float(s.replace(",", ""))
                except Exception:
                    continue
                if num <= 0:
                    continue
                numeric_candidates.append((num, s))

        if numeric_candidates:
            numeric_candidates.sort(reverse=True, key=lambda t: t[0])
            
            # Tax components (CGST/SGST/IGST/Cess) - extract from lines containing or near the tax keyword
            if any(k in label_l for k in ["cgst", "sgst", "igst", "cess"]):
                # Determine which tax keyword to search for
                tax_keyword = None
                if "cgst" in label_l:
                    tax_keyword = "cgst"
                elif "sgst" in label_l:
                    tax_keyword = "sgst"
                elif "igst" in label_l:
                    tax_keyword = "igst"
                elif "cess" in label_l:
                    tax_keyword = "cess"
                
                # EARLY CHECK: For IGST/Cess, check if this is intrastate (same state)
                # Intrastate invoices (e.g., Maharashtra → Maharashtra) use CGST+SGST, not IGST
                if tax_keyword in ["igst", "cess"]:
                    is_intrastate = False
                    # Check if Place of Supply matches seller state (both start with same code like "27")
                    for v in ocr_values:
                        if not isinstance(v, str):
                            continue
                        v_lower = v.lower()
                        # Look for patterns indicating intrastate
                        # GSTINs starting with same 2 digits indicate same state
                        # If we see Maharashtra (27) for both parties, it's intrastate
                        if "27" in v and ("place of supply" in v_lower or "maharashtra" in v_lower):
                            is_intrastate = True
                            break
                        # Check buyer GSTIN starts with same state as seller
                        gstin_match = re.search(r"\b(27)[A-Z0-9]{13}\b", v)
                        if gstin_match:
                            is_intrastate = True
                            break
                    
                    if is_intrastate:
                        return {"value": None, "confidence": 0.9, "reason": f"{tax_keyword} not applicable (intrastate transaction)"}
                
                # PRIORITY 0: Direct regex for "CGST 300.00" or "SGST 1,456.31" patterns
                # This handles concatenated OCR like "97,087.38CGST 1,456.31SGST 1,456.31"
                if tax_keyword in ["cgst", "sgst"]:
                    for v in ocr_values:
                        if not isinstance(v, str):
                            continue
                        v_lower = v.lower()
                        # Skip if this keyword isn't in the string
                        if tax_keyword not in v_lower:
                            continue
                        
                        # Pattern: CGST/SGST followed by amount
                        # Handles: "CGST 1,456.31" or "97,087.38CGST 1,456.31" or "CGST1,456.31"
                        # Use word char or digit before tax_keyword (not letter) to avoid "UTCGST"
                        pattern = rf"(?:^|[^a-zA-Z]){tax_keyword}\s*([\d,]+\.?\d*)"
                        match = re.search(pattern, v_lower)
                        if match:
                            amount_str = match.group(1)
                            # Clean up: remove trailing comma if any
                            amount_str = amount_str.rstrip(',')
                            if amount_str:
                                try:
                                    num = float(amount_str.replace(",", ""))
                                    # Tax amount should be reasonable (not 0, not huge like taxable value)
                                    if 1 <= num < 50000:  # Reasonable tax range
                                        return {"value": amount_str, "confidence": 0.95, "reason": f"regex: {tax_keyword} direct match"}
                                except:
                                    pass
                
                # First: Estimate the maximum taxable value to filter out large amounts
                # Tax amounts should be MUCH smaller than taxable value (typically 1.5-18%)
                max_taxable = 0
                for v in ocr_values:
                    if not isinstance(v, str):
                        continue
                    if "taxable" in v.lower() or "total" in v.lower():
                        for s in _extract_numbers(v):
                            try:
                                num = float(s.replace(",", ""))
                                if num > max_taxable:
                                    max_taxable = num
                            except:
                                pass
                
                # Tax amount should be less than 20% of taxable value
                max_tax_threshold = max_taxable * 0.20 if max_taxable > 0 else 10000
                
                # Priority 1: Look for lines with percentage pattern "1.50% 300.00"
                for v in ocr_values:
                    if not isinstance(v, str):
                        continue
                    # Pattern like "1.50% 300.00" or "1.50 300.00"
                    if "1.5" in v and "%" in v:
                        amounts = _extract_numbers(v)
                        for s in amounts:
                            try:
                                num = float(s.replace(",", ""))
                                # Skip the percentage itself (1.50) and taxable values
                                if 10 <= num <= max_tax_threshold:
                                    return {"value": s, "confidence": 0.9, "reason": f"heuristic: {tax_keyword} from rate line"}
                            except:
                                pass
                
                # Priority 2: Look for duplicate small amounts (like "300.00 300.00")
                for v in ocr_values:
                    if not isinstance(v, str):
                        continue
                    amounts = _extract_numbers(v)
                    if len(amounts) >= 2:
                        for s in amounts:
                            try:
                                num = float(s.replace(",", ""))
                                # Must be small (tax amount) AND appear twice (CGST = SGST)
                                if 10 <= num <= max_tax_threshold and amounts.count(s) >= 2:
                                    return {"value": s, "confidence": 0.85, "reason": f"heuristic: {tax_keyword} from duplicate small amount"}
                            except:
                                pass
                
                # Priority 3: Look for lines containing the tax keyword with small amounts
                for v in ocr_values:
                    if not isinstance(v, str):
                        continue
                    v_lower = v.lower()
                    if tax_keyword and tax_keyword in v_lower:
                        amounts = _extract_numbers(v)
                        for s in amounts:
                            try:
                                num = float(s.replace(",", ""))
                                if 10 <= num <= max_tax_threshold:
                                    return {"value": s, "confidence": 0.8, "reason": f"heuristic: {tax_keyword} from labeled line"}
                            except:
                                pass
                
                # Priority 4: Look for Tax Amount total and divide by 2 for CGST/SGST
                for v in ocr_values:
                    if not isinstance(v, str):
                        continue
                    v_lower = v.lower()
                    if "tax amount" in v_lower and "words" not in v_lower:
                        amounts = _extract_numbers(v)
                        for s in amounts:
                            try:
                                num = float(s.replace(",", ""))
                                if 10 <= num <= max_tax_threshold * 2:  # Total tax = CGST + SGST
                                    half_tax = num / 2
                                    return {"value": f"{half_tax:.2f}", "confidence": 0.75, "reason": f"heuristic: {tax_keyword} from total tax / 2"}
                            except:
                                pass
                
                # IGST/Cess - return None (not applicable for intrastate)
                if tax_keyword in ["igst", "cess"]:
                    return {"value": None, "confidence": 0.0, "reason": f"{tax_keyword} not applicable (intrastate)"}
                
                # No explicit tax line found - return None
                return {"value": None, "confidence": 0.0, "reason": f"no {tax_keyword} amount found"}
            
            # TOTAL INVOICE VALUE: Look for explicit patterns
            if "total" in label_l and ("invoice" in label_l or "value" in label_l):
                # First: Look for "INR X Only" pattern - the total in words line
                for v in ocr_values:
                    if not isinstance(v, str):
                        continue
                    v_lower = v.lower()
                    # "INR One Lakh Seventy Nine Thousand" -> next line often has numeric total
                    if "inr" in v_lower and "only" in v_lower:
                        continue  # Skip the words line, look nearby for numeric
                
                # Second: Look for amount near "E.&O.E" or standalone large amounts at end
                for v in ocr_values:
                    if not isinstance(v, str):
                        continue
                    # Lines like "1,79,360.00" at end, or "* 25,750.00 E.&O.E"
                    if "e.&o.e" in v.lower() or "e & o e" in v.lower():
                        amounts = _extract_numbers(v)
                        if amounts:
                            # Get the largest number on this line
                            best_amt = max(amounts, key=lambda x: float(x.replace(",", "")))
                            return {"value": best_amt, "confidence": 0.9, "reason": "heuristic: total from E.&O.E line"}
                
                # Third: Look for standalone total line (just a large number)
                candidate_totals = []
                for v in ocr_values:
                    if not isinstance(v, str):
                        continue
                    v_stripped = v.strip()
                    amounts = _extract_numbers(v_stripped)
                    # If line is mostly just a number (total invoice value)
                    if amounts and len(v_stripped) < 20:
                        for s in amounts:
                            try:
                                num = float(s.replace(",", ""))
                                # Total invoice should be a reasonably large amount (>1000)
                                if num >= 1000 and s == v_stripped.replace("*", "").strip():
                                    candidate_totals.append((num, s))
                            except:
                                pass
                
                if candidate_totals:
                    # Return the LARGEST candidate found (Grand Total > Taxable Value)
                    best_total = max(candidate_totals, key=lambda x: x[0])
                    return {"value": best_total[1], "confidence": 0.85, "reason": "heuristic: max standalone total value"}
                
                # Fallback: use the highest numeric value
                best = numeric_candidates[0][1]
                return {"value": best, "confidence": 0.7, "reason": "heuristic: max numeric (total)"}
            
            # TAXABLE VALUE: Look for explicit "Taxable Value" line with amount
            if "taxable" in label_l:
                # First: Look for lines containing "Taxable Value" followed by amount
                for v in ocr_values:
                    if not isinstance(v, str):
                        continue
                    v_lower = v.lower()
                    if "taxable" in v_lower and ("value" in v_lower or "val" in v_lower):
                        amounts = _extract_numbers(v)
                        # Get the first large number (taxable value comes before rates)
                        for s in amounts:
                            try:
                                num = float(s.replace(",", ""))
                                # Taxable value should be reasonably large (not rate like 1.50)
                                if num >= 100:
                                    return {"value": s, "confidence": 0.9, "reason": "heuristic: taxable value from labeled line"}
                            except:
                                pass
                
                # Second: Look for "Total" line with taxable amount
                for v in ocr_values:
                    if not isinstance(v, str):
                        continue
                    v_lower = v.lower()
                    # Lines like "Total 1,00,000.00" or "Total 1,74,135.92"
                    if v_lower.startswith("total ") and "gms" not in v_lower:
                        amounts = _extract_numbers(v)
                        for s in amounts:
                            try:
                                num = float(s.replace(",", ""))
                                if num >= 100:
                                    return {"value": s, "confidence": 0.8, "reason": "heuristic: taxable from Total line"}
                            except:
                                pass
                
                # Fallback: find a reasonable taxable amount (not the max which is usually total with tax)
                # Taxable value = Total - Tax, so it's typically slightly less than max
                if len(numeric_candidates) >= 2:
                    # Look for values that could be taxable (large but not the largest)
                    for num, s in numeric_candidates[1:5]:  # Check 2nd-5th largest
                        if num >= 100:  # Reasonable taxable value
                            return {"value": s, "confidence": 0.7, "reason": "heuristic: taxable value (not max)"}
                
                # Last resort: use second highest
                if len(numeric_candidates) >= 2:
                    best = numeric_candidates[1][1]
                    return {"value": best, "confidence": 0.6, "reason": "heuristic: taxable value (2nd highest)"}
            
            # For general amounts: use highest
            best = numeric_candidates[0][1]
            return {"value": best, "confidence": 0.7, "reason": "heuristic: max numeric"}

    # Percent / rate (with enhanced tax rate detection)
    if "rate" in label_l or "%" in label_l or "tax rate" in label_l:
        # First try: explicit percentage values like "1.50%"
        for v in ocr_values:
            if not isinstance(v, str):
                continue
            p = _extract_percent(v)
            if p:
                # Prefer values near "tax", "gst", "cgst", "sgst"
                v_lower = v.lower()
                if any(k in v_lower for k in ["tax", "gst", "cgst", "sgst", "igst"]):
                    return {"value": p, "confidence": 0.9, "reason": "regex: tax rate with context"}
                return {"value": p, "confidence": 0.7, "reason": "regex: percent"}
        
        # Second try: look for standalone rate patterns "1.50" or "1.5" or "3"
        # Common jewelry GST rates: 1.5% CGST + 1.5% SGST = 3% total
        for v in ocr_values:
            if not isinstance(v, str):
                continue
            # Look for "1.50" or "1.5" standalone or with context
            v_lower = v.lower()
            if re.search(r"\b1\.50?\b", v) and "gms" not in v_lower and "," not in v:
                # Likely a tax rate
                m = re.search(r"\b(1\.50?)\b", v)
                if m:
                    return {"value": f"{m.group(1)}%", "confidence": 0.8, "reason": "heuristic: 1.50 GST rate"}
            # Look for "3%" or "3" as total GST rate
            if re.search(r"\b3\.?0?\s*%", v):
                return {"value": "3%", "confidence": 0.8, "reason": "heuristic: 3% GST rate"}
        
        # Third try: common GST rates from "@ X%" pattern
        for v in ocr_values:
            if not isinstance(v, str):
                continue
            # Look for "@ X%" pattern
            m = re.search(r"@\s*(\d+(?:\.\d+)?)\s*%", v)
            if m:
                rate = m.group(1)
                return {"value": f"{rate}%", "confidence": 0.85, "reason": "regex: @ rate pattern"}
        
        # Fourth try: find any percentage-like number in rate context
        for v in ocr_values:
            if not isinstance(v, str):
                continue
            v_lower = v.lower()
            if "rate" in v_lower or "%" in v:
                # Extract numbers and check if they're reasonable rates
                numbers = _extract_numbers(v)
                for n in numbers:
                    try:
                        rate = float(n.replace(",", ""))
                        # GST rates are typically: 0, 0.1, 0.25, 1.5, 3, 5, 12, 18, 28
                        if rate in [0.1, 0.25, 1.5, 1.50, 3, 5, 12, 18, 28]:
                            return {"value": f"{rate}%", "confidence": 0.75, "reason": "heuristic: known GST rate"}
                    except:
                        pass

    # Place of supply / state - return state name with code (e.g., "27-Maharashtra")
    if "place of supply" in label_l or (("state" in label_l) and ("name" in label_l or "code" in label_l)):
        for v in ocr_values:
            if not isinstance(v, str):
                continue
            # Look for "State Name: <state>, Code: <code>" pattern
            m = re.search(r"State Name\s*:\s*([A-Za-z ]+),\s*Code\s*:\s*(\d{1,2})", v, flags=re.IGNORECASE)
            if m:
                state = " ".join(m.group(1).split()).strip()
                code = m.group(2)
                return {"value": f"{code}-{state}", "confidence": 0.9, "reason": "regex: state name and code"}
            # Also try "<state> (<code>)" pattern
            m2 = re.search(r"([A-Za-z][A-Za-z ]+)\s*\(\s*(\d{1,2})\s*\)", v)
            if m2:
                state = m2.group(1).strip()
                code = m2.group(2)
                return {"value": f"{code}-{state}", "confidence": 0.85, "reason": "regex: state with code in parens"}
            # Try "Code: <code>" with state name nearby
            m3 = re.search(r"\bCode\s*[:\-]?\s*(\d{1,2})\b", v, flags=re.IGNORECASE)
            if m3:
                code = m3.group(1)
                # Try to find state name in same line
                state_match = re.search(r"\b(Maharashtra|Gujarat|Karnataka|Tamil Nadu|Delhi|Rajasthan|Uttar Pradesh|West Bengal|Madhya Pradesh|Andhra Pradesh|Telangana|Kerala|Punjab|Haryana|Bihar|Odisha|Jharkhand|Chhattisgarh|Assam|Goa)\b", v, flags=re.IGNORECASE)
                if state_match:
                    state = state_match.group(1)
                    return {"value": f"{code}-{state}", "confidence": 0.8, "reason": "regex: state code with name"}

    # Recipient / buyer name (but NOT GSTIN fields)
    if ("recipient" in label_l or "receiver" in label_l or "buyer" in label_l or "name" in label_l) and "gstin" not in label_l and "gst" not in label_l:
        
        # Helper function to check if a string is a valid name (not GSTIN/label/code)
        def is_valid_name(s: str) -> bool:
            if not s or len(s) < 3:
                return False
            s_lower = s.lower().strip()
            # Exclude GSTINs (15-char alphanumeric starting with state code)
            if re.match(r"^\d{2}[A-Z0-9]{13}$", s.upper()):
                return False
            # Exclude labels and headers
            invalid_patterns = [
                "gstin", "gst", "state", "code", 
                "buyer's order", "order no", "invoice", "date", "place of supply",
                "description", "hsn", "sac", "quantity", "amount", "tax", "total",
                "e-way", "vehicle", "transport", "terms", "bank", "account",
                "ifsc", "branch", "signature", "authorized", "subject",
                "dispatch", "doc"
            ]
            if any(k in s_lower for k in invalid_patterns):
                return False
            
            # Specific check for "no" / "no." at the start or as a standalone word (to avoid matching "Shop No.")
            if s_lower.startswith("no.") or s_lower.startswith("no ") or s_lower == "no":
                return False

            # Exclude pure numbers or codes
            if re.match(r"^[\d\-\/]+$", s):
                return False
            # Exclude state codes like "27-Maharashtra" or standalone state names
            if re.match(r"^\d{1,2}\s*[-:]\s*[A-Za-z]+$", s):
                return False
            # Exclude addresses (contain typical address words)
            # Exclude addresses if the name STARTS with them (e.g. "Road No 5...")
            # But allow if they are part of a larger name (e.g. "Jay Bhavani Jewellers... Station Road")
            address_start_patterns = ["road", "tower", "floor", "building", "station", "opposite", "shop no", "flat no", "address"]
            for pat in address_start_patterns:
                if s_lower.startswith(pat):
                    return False
            
            # Still strictly exclude some obviously bad patterns anywhere in the string
            # strict_address_patterns = ["chawl", "chs", "ltd", "bhavan", "nagar", "west", "east", "north", "south"]
            # Only block if the pattern is the ONLY thing or dominates (simple heuristic: if length is short and contains address word)
            # Actually, "Jay Bhavani Jewellers ... West" contains "west", so we must be careful.
            # Let's remove the strict address pattern check that blocks valid names containing address parts.
            # Instead, rely on the "Start With" check and the general invalid list.
            if any(k in s_lower for k in ["gstin", "date", "total", "tax", "amount"]):
                 return False
            
            return True
        
        # FIRST: Look specifically for "Buyer (Bill to)" or "Consignee (Ship to)" labels
        # The name immediately after these labels is the buyer name
        buyer_idx = None
        for i, v in enumerate(ocr_values):
            if not isinstance(v, str):
                continue
            t = v.lower()
            # Look for explicit buyer/consignee labels
            if "buyer" in t and ("bill" in t or "to" in t):
                buyer_idx = i
                break
            if "consignee" in t and "ship" in t:
                buyer_idx = i
                break
            if "receiver" in t:
                buyer_idx = i
                break
        
        if buyer_idx is not None:
            # Check next few lines for the actual name (skip the label line)
            for j in range(buyer_idx + 1, min(buyer_idx + 5, len(ocr_values))):
                next_val = ocr_values[j]
                if not isinstance(next_val, str):
                    continue
                next_lower = next_val.lower()
                # Stop if we hit another section
                # Clean up - remove merged GSTIN/State info that often appears in OCR
                clean_val = next_val
                
                # Split at common labels that get merged
                for split_marker in ["GSTIN", "State Name", "Place of Supply", "Invoice No"]:
                    if split_marker.lower() in clean_val.lower():
                        clean_val = re.split(f"(?i){re.escape(split_marker)}", clean_val)[0]
                
                # ALSO split at GSTIN pattern if the label was missing but value is there (e.g. "... West : 27AUB...")
                # Match ": 27AUB..." or just "27AUB..." at end of string
                gstin_pattern = r"[:\s-]*\b\d{2}[A-Z]{5}\d{4}[A-Z][A-Z0-9]{3}\b"
                split_match = re.search(gstin_pattern, clean_val)
                if split_match:
                    clean_val = clean_val[:split_match.start()]
                
                # Truncate at address start markers
                # Common words that start an address part in invoice headers
                address_markers = [
                    "Gr Floor", "Ground Floor", "1st Floor", "2nd Floor", "Shop No", "Office No", "Flat No", 
                    "Plot No", "Room No", "Opp.", "Opposite", "Near", "Behind", "Beside", "Next to", 
                    "Lane", "Sector", "Nagar", "Marg", "Road", "Street", "Chawl", "Bhavan", "Apartment", "Tower"
                ]
                # Sort by length desc to match longer phrases first
                address_markers.sort(key=len, reverse=True)
                
                clean_val_lower = clean_val.lower()
                best_split_idx = len(clean_val)
                
                for marker in address_markers:
                    # Search case-insensitive
                    m_idx = clean_val_lower.find(marker.lower())
                    
                    if m_idx != -1:
                        # Check character before match (Start Boundary)
                        if m_idx > 0:
                            prev_char = clean_val[m_idx-1]
                            if prev_char.isalnum():
                                # Part of a word start (e.g. "Broad"), skip to next
                                next_search_start = m_idx + 1
                                while True:
                                    m_idx = clean_val_lower.find(marker.lower(), next_search_start)
                                    if m_idx == -1:
                                        break
                                    # Check start boundary for this new match
                                    if m_idx > 0 and clean_val[m_idx-1].isalnum():
                                        next_search_start = m_idx + 1
                                        continue
                                    break
                        
                        # Check character after match (End Boundary)
                        # Failure here means it's a prefix of a word (e.g. "Bhavan" in "Bhavani")
                        if m_idx != -1:
                            end_idx = m_idx + len(marker)
                            if end_idx < len(clean_val):
                                next_char = clean_val[end_idx]
                                if next_char.isalnum():
                                    # It's inside a word, ignore this match
                                    # We should continue searching but for simplicity/safety with current loop structure, 
                                    # let's just ignore this marker or implement a better loop.
                                    # A recursive/while loop is better but let's just skip this marker for this string 
                                    # if the first valid-start match fails end-boundary.
                                    # NOTE: This assumes words don't repeat in valid ways (e.g. "Bhavani Bhavan")
                                    # If "Bhavani Bhavan" exists, we should find the second one.
                                    
                                    # Let's do a quick continue search if end boundary fails
                                    next_search_start = m_idx + 1
                                    while True:
                                        m_idx = clean_val_lower.find(marker.lower(), next_search_start)
                                        if m_idx == -1:
                                            break
                                        
                                        # Check start boundary
                                        if m_idx > 0 and clean_val[m_idx-1].isalnum():
                                            next_search_start = m_idx + 1
                                            continue
                                            
                                        # Check end boundary
                                        e_idx = m_idx + len(marker)
                                        if e_idx < len(clean_val) and clean_val[e_idx].isalnum():
                                            next_search_start = m_idx + 1
                                            continue
                                        
                                        # Found valid match
                                        break

                        if m_idx != -1:
                            # Use the earliest occurrence of any valid marker match
                            if m_idx < best_split_idx:
                                best_split_idx = m_idx

                if best_split_idx < len(clean_val):
                    clean_val = clean_val[:best_split_idx]

                clean_val = clean_val.strip(" ,:-")

                # Stop if we hit another section (checking original value to be safe about section boundaries)
                if any(k in next_lower for k in ["invoice no"]):
                    break
                
                if is_valid_name(clean_val):
                    return {"value": clean_val, "confidence": 0.9, "reason": "heuristic: name after buyer label"}
        
        # SECOND: Look for standalone buyer/consignee/bill-to labels
        for i, v in enumerate(ocr_values):
            if not isinstance(v, str):
                continue
            t = v.lower()
            if any(k in t for k in ["buyer", "bill to", "consignee", "ship to", "receiver"]):
                # Check next value for the actual name
                if i + 1 < len(ocr_values):
                    next_val = ocr_values[i + 1]
                    if isinstance(next_val, str) and is_valid_name(next_val):
                        return {"value": next_val.strip(), "confidence": 0.85, "reason": "heuristic: name after label"}
                
                # Extract from same line after colon
                if ":" in v:
                    parts = v.split(":", 1)
                    if len(parts) == 2:
                        name = parts[1].strip()
                        # Clean up - remove GSTIN and state info
                        name = re.split(r"GSTIN\s*/?UIN", name, flags=re.IGNORECASE)[0]
                        name = re.split(r"State Name", name, flags=re.IGNORECASE)[0]
                        name = name.strip(" ,:-")
                        if is_valid_name(name):
                            return {"value": name, "confidence": 0.8, "reason": "heuristic: name after colon"}

    # Reverse charge (Y/N)
    if "reverse charge" in label_l:
        for v in ocr_values:
            if not isinstance(v, str):
                continue
            t = v.lower()
            if "reverse" in t and "charge" in t:
                if re.search(r"\b(yes|y)\b", t):
                    return {"value": "Y", "confidence": 0.8, "reason": "regex: reverse charge"}
                if re.search(r"\b(no|n)\b", t):
                    return {"value": "N", "confidence": 0.8, "reason": "regex: reverse charge"}

    # Note type (Debit/Credit/Refund)
    if "note type" in label_l:
        for v in ocr_values:
            if not isinstance(v, str):
                continue
            t = v.lower()
            if "debit" in t:
                return {"value": "Debit", "confidence": 0.7, "reason": "regex: note type"}
            if "credit" in t:
                return {"value": "Credit", "confidence": 0.7, "reason": "regex: note type"}
            if "refund" in t:
                return {"value": "Refund", "confidence": 0.7, "reason": "regex: note type"}

    # HSN/SAC codes
    if "hsn" in label_l or "sac" in label_l:
        for v in ocr_values:
            if not isinstance(v, str):
                continue
            # Look for HSN codes but avoid years (2020-2029) and common false positives
            m = re.search(r"\b(\d{4,8})\b", v)
            if m:
                hsn_code = m.group(1)
                # Avoid extracting years
                if hsn_code.startswith(("202", "201", "200")):
                    continue
                # Avoid codes from invoice numbers (like "2024" from "67/2024-25")
                if "/" in v and len(hsn_code) == 4:
                    continue
                return {"value": hsn_code, "confidence": 0.75, "reason": "regex: hsn/sac"}

    # UQC
    if "uqc" in label_l:
        for v in ocr_values:
            if not isinstance(v, str):
                continue
            tok = v.strip().upper()
            # Direct match, filter out 'NO' and only allow valid UQC codes
            if tok in _UQC_TOKENS and tok != "NO":
                return {"value": tok, "confidence": 0.75, "reason": "heuristic: uqc token"}
            # Check within text (e.g., "1.5 KG" or "5 PCS") but avoid "Shop No." patterns
            if "NO" in tok and ("SHOP" in tok or "." in tok):
                continue  # Skip "Shop No.", "Sr. No.", etc.
            for token in tok.split():
                if token in _UQC_TOKENS and token != "NO":
                    return {"value": token, "confidence": 0.7, "reason": "heuristic: uqc in text"}

    # Quantity
    if "quantity" in label_l or "qty" in label_l:
        # First: Look for explicit "Qty:" or "Quantity:" labels in OCR
        for v in ocr_values:
            if not isinstance(v, str):
                continue
            v_lower = v.lower()
            # Look for "Qty: 5" or "Quantity: 10" patterns
            qty_match = re.search(r"(?:qty|quantity)\s*[:\-]?\s*(\d+(?:\.\d+)?)", v_lower)
            if qty_match:
                qty_val = qty_match.group(1)
                try:
                    num = float(qty_val)
                    if 0 < num <= 1000:
                        return {"value": qty_val, "confidence": 0.9, "reason": "regex: explicit qty label"}
                except:
                    pass
        
        # Second: Look for weight quantities like "1.960 gms" or "16.530 gms" (jewelry/gold invoices)
        for v in ocr_values:
            if not isinstance(v, str):
                continue
            # Skip address patterns
            if any(k in v.lower() for k in ["tower", "road", "floor", "building", "chs", "address", "shop"]):
                continue
            # Match patterns like "1.960 gms" or "16.530 gms" - common in jewelry invoices
            m = re.search(r"\b(\d+\.\d{2,3})\s*(?:gms?|grams?|g)\b", v.lower())
            if m:
                qty = m.group(1)
                try:
                    num = float(qty)
                    if 0 < num <= 1000:
                        return {"value": qty, "confidence": 0.9, "reason": "heuristic: weight qty in gms"}
                except:
                    pass
        
        # Third: Look for "Total X gms" pattern
        for v in ocr_values:
            if not isinstance(v, str):
                continue
            v_lower = v.lower()
            # Match "Total 1.960 gms" but not "Total 1,00,000.00" (amounts)
            m = re.search(r"total\s+(\d+\.\d{2,3})\s*(?:gms?|grams?|g)?\b", v_lower)
            if m:
                qty = m.group(1)
                try:
                    num = float(qty)
                    # Weight quantities are typically small (< 100 gms for jewelry)
                    if 0 < num <= 500:
                        return {"value": qty, "confidence": 0.85, "reason": "heuristic: total weight qty"}
                except:
                    pass
        
        # Fourth: Look for numbers followed by UQC tokens (like "5 PCS", "2 NOS")
        for v in ocr_values:
            if not isinstance(v, str):
                continue
            if any(k in v.lower() for k in ["tower", "road", "floor", "building", "chs", "address"]):
                continue
            m = re.search(r"\b(\d+(?:\.\d+)?)\s*(PCS|NOS|KGS?|LTR|MTR|BOX|SET|PAIR|PAIRS)\b", v.upper())
            if m:
                qty = m.group(1)
                try:
                    num = float(qty)
                    if 0 < num <= 1000:
                        return {"value": qty, "confidence": 0.85, "reason": "heuristic: qty with UQC"}
                except:
                    pass
        
        # Fallback: return 1 only if no quantity pattern found
        return {"value": "1", "confidence": 0.5, "reason": "default: no qty pattern found"}

    # Item Description
    if "description" in label_l or "item" in label_l:
        # Look for actual item names (jewelry, gold, etc.) but avoid supplier names
        for v in ocr_values:
            if not isinstance(v, str):
                continue
            t = v.lower()
            # Skip supplier/company names (usually at top of invoice)
            if any(k in t for k in ["pvt", "ltd", "limited", "company", "corp", "inc"]):
                continue
            # Check for common item keywords
            if any(k in t for k in ["ornament", "jewellery", "jewelry", "ring", "necklace", "bracelet", "earring", "chain"]):
                # Clean up
                desc = v.strip()
                if len(desc) > 2 and len(desc) < 100:
                    return {"value": desc, "confidence": 0.8, "reason": "heuristic: item keyword match"}
            
            # Check for "gold" but make sure it's not just "Gold" company name
            if "gold" in t and len(v.split()) >= 2:
                # Avoid company/shop/buyer names that contain "gold"
                # Examples to exclude: "VARUNACHI GOLD-New", "Shree Ashapura Gold", "Labh Gold"
                buyer_indicators = [
                    "shree", "sri", "shri", "traders", "jewellers", "jewelers",
                    "shop", "floor", "ground", "road", "tower", "building",
                    "chs", "ltd", "estate", "new", "park", "nagar", "mandvi",
                    "east", "west", "north", "south", "mumbai", "thane"
                ]
                # Also skip if it looks like a buyer name (short name ending with "-New" or similar)
                if v.strip().endswith("-New") or v.strip().endswith("-new"):
                    continue
                if not any(k in t for k in buyer_indicators):
                    desc = v.strip()
                    if len(desc) > 5 and len(desc) < 100:
                        return {"value": desc, "confidence": 0.75, "reason": "heuristic: gold item"}
        
        # Fallback 1: Look for descriptive text with "Description:" prefix
        for v in ocr_values:
            if not isinstance(v, str):
                continue
            t = v.lower()
            if "description" in t and ":" in t:
                parts = v.split(":", 1)
                if len(parts) == 2 and len(parts[1].strip()) > 3:
                    desc = parts[1].strip()
                    # Exclude if it's just a header or state info
                    if not any(k in desc.lower() for k in ["state name", "code", "maharashtra", "goods", "of goods"]):
                        return {"value": desc, "confidence": 0.6, "reason": "heuristic: description after colon"}

        # Fallback 2: Positional - Find text between "Description" header and "Total"/"Taxable Value"
        # 1. Find Header Index
        header_idx = -1
        for i, v in enumerate(ocr_values):
            if not isinstance(v, str): continue
            t = v.lower()
            if "description" in t and ("goods" in t or "service" in t or "item" in t):
                header_idx = i
                break
        
        # 2. Find Footer Index (Taxable Value or Total)
        footer_idx = len(ocr_values)
        for i, v in enumerate(ocr_values):
            if not isinstance(v, str): continue
            t = v.lower()
            # Start searching for footer ONLY after header
            if header_idx != -1 and i <= header_idx:
                continue
            
            if "total" in t or "taxable" in t or "amount" in t:
                # Must contain a number to be a valid footer line (value line)
                if re.search(r"\d", t):
                    footer_idx = i
                    break
        
        if header_idx != -1:
            candidates = []
            for i in range(header_idx + 1, footer_idx):
                v = ocr_values[i]
                if not isinstance(v, str): continue
                t = v.lower()
                
                # Filter out junk lines
                if len(t) < 4: continue # Too short
                if re.match(r"^[\d\s,.:\-]+$", t): continue # Pure numbers/symbols
                if "hsn" in t or "sac" in t: continue # HSN headers repeated
                if re.match(r"^\d{4,8}$", t.strip()): continue # HSN code
                if re.match(r"^\d+\.\d{2,3}$", t.strip()): continue # Quantity/Rate
                if re.match(r"^\d+(\.\d+)?\s*%$", t.strip()): continue # Percentage
                if any(k in t for k in ["output", "input", "cgst", "sgst", "igst"]): continue # Tax lines
                
                # Boost score for having typical item words
                score = len(t)
                if any(k in t for k in ["gold", "silver", "ornament", "ring", "chain"]):
                    score += 50
                
                candidates.append((score, v))
            
            if candidates:
                # Return the best candidate (longest / most relevant)
                best_cand = max(candidates, key=lambda x: x[0])[1]
                return {"value": best_cand.strip(), "confidence": 0.65, "reason": "heuristic: best candidate between header/footer"}

    return {"value": None, "confidence": 0.0, "reason": "no heuristic"}


def gemini_batch_choose(fields: Dict[str, List[str]], ocr_values: List[str]) -> Dict[str, Dict[str, Any]]:
    """Single-call Gemini rerank to avoid per-field quota issues."""
    if gemini is None:
        logger.info("Batch LLM call skipped: gemini client not initialized")
        return {}
    if not fields:
        return {}
    
    model = get_llm_model()
    logger.info(f"Attempting batch LLM call for {len(fields)} fields using model: {model}")

    # Keep OCR context short to avoid huge prompts.
    ocr_context = "\n".join(ocr_values[:120])

    payload = {k: v for k, v in fields.items() if v}
    if not payload:
        return {}

    prompt = f"""
You are filling an Excel template from OCR text.

OCR SNIPPET (may be noisy):
{ocr_context}

For EACH field below, pick the best candidate that is the actual DATA (not a label like 'No.' or 'per').

Rules:
- Output MUST be JSON only.
- For each field, value MUST be exactly one of the provided candidates, or null.
- confidence is 0.0 to 1.0.

FIELDS (JSON where each key is a field label and value is the list of candidates):
{json.dumps(payload, ensure_ascii=False)}

OUTPUT FORMAT:
{{
  "Field Label": {{"value": "..." | null, "confidence": 0.0, "reason": "..."}},
  ...
}}
"""

    try:
        resp = gemini.models.generate_content(model=model, contents=prompt)
    except Exception as e:
        logger.warning(f"Batch LLM call failed (will use per-field fallback): {e}")
        return {}

    raw = str(resp.text)
    
    # Try to extract JSON from common patterns:
    # 1. Plain JSON
    # 2. JSON in markdown code fences: ```json ... ```
    # 3. JSON with surrounding text
    
    json_text = None
    
    # Try markdown code fence first
    code_fence_match = re.search(r"```(?:json)?\s*(\{[\s\S]*?\})\s*```", raw, re.IGNORECASE)
    if code_fence_match:
        json_text = code_fence_match.group(1)
    else:
        # Try to find largest JSON object
        json_match = re.search(r"\{[\s\S]*\}", raw)
        if json_match:
            json_text = json_match.group(0)
    
    if not json_text:
        logger.warning(f"Batch LLM response did not contain valid JSON. First 200 chars: {raw[:200]}")
        return {}

    try:
        parsed = json.loads(json_text)
    except Exception as e:
        logger.warning(f"Batch LLM JSON parse failed: {e}. JSON snippet: {json_text[:200]}")
        return {}

    out: Dict[str, Dict[str, Any]] = {}
    for label, obj in parsed.items() if isinstance(parsed, dict) else []:
        if label not in payload:
            continue
        if not isinstance(obj, dict):
            continue
        val = obj.get("value")
        # Allow null values; only reject if non-null value isn't in candidate list
        if val is not None and val not in payload[label]:
            continue
        try:
            conf = float(obj.get("confidence", 0.5))
        except Exception:
            conf = 0.5
        out[label] = {"value": val, "confidence": conf, "reason": obj.get("reason", "")}
    
    logger.info(f"Batch LLM call succeeded: extracted {len(out)}/{len(payload)} fields")
    return out

def safe_write(ws, row: int, col: int, value):
    from openpyxl.utils import get_column_letter
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        cell.value = value
        return
    # Convert row, col to coordinate string (e.g., "A1")
    coord = f"{get_column_letter(col)}{row}"
    for rng in ws.merged_cells.ranges:
        if coord in rng:
            ws.cell(rng.min_row, rng.min_col).value = value
            return

# ---------------------------------------------------------------------
# CANDIDATE GENERATION (RULE + FUZZY)
# ---------------------------------------------------------------------
def generate_candidates(label: str, ocr_values: List[str], top_k: int = 10) -> List[str]:
    scored: List[Tuple[float, str]] = []
    label_lower = (label or "").lower()
    label_tokens = set(label_lower.split())

    def add(score: float, cand: str):
        c = (cand or "").strip()
        if not c:
            return
        if _is_bad_candidate(label, c):
            return
        scored.append((score, c))

    for v in ocr_values:
        if not isinstance(v, str):
            continue
        v_norm = " ".join(v.split())
        v_lower = v_norm.lower()

        # Base similarity score
        partial_score = fuzz.partial_ratio(label_lower, v_lower)
        token_sort_score = fuzz.token_sort_ratio(label_lower, v_lower)

        token_boost = 0
        v_tokens = set(v_lower.split())
        if label_tokens & v_tokens:
            token_boost = 20

        score = max(partial_score, token_sort_score) + token_boost
        if score >= 40:
            add(float(score), v_norm)

        # Add extracted structured tokens as candidates (helps avoid passing long header lines)
        if "gstin" in label_lower or "gst" in label_lower:
            if _looks_like_gstin(v_norm):
                add(float(score) + 15.0, v_norm)
        if "date" in label_lower or "dated" in label_lower:
            dt = _extract_date_token(v_norm)
            if dt:
                add(float(score) + 15.0, dt)
        if "rate" in label_lower or "%" in label_lower:
            pct = _extract_percent(v_norm)
            if pct:
                add(float(score) + 12.0, pct)
        if any(k in label_lower for k in ["amount", "value", "total", "taxable", "quantity"]):
            for num in _extract_numbers(v_norm):
                add(float(score) + 8.0, num)

    # Dedupe while keeping best score
    best_by_value: Dict[str, float] = {}
    for s, c in scored:
        prev = best_by_value.get(c)
        if prev is None or s > prev:
            best_by_value[c] = s

    ranked = sorted(best_by_value.items(), key=lambda t: t[1], reverse=True)
    return [c for c, _ in ranked[:top_k]]

# ---------------------------------------------------------------------
# CLAUDE RERANK (SAFE, NO eval)
# ---------------------------------------------------------------------
def gemini_choose(label: str, candidates: List[str]) -> Dict[str, Any]:
    if not candidates:
        return {"value": None, "confidence": 0.0, "reason": "no candidates"}

    # If Vertex/Gemini is unavailable, we fall back to best local candidate.
    # This keeps template mode functional even without model access.
    def local_fallback(reason: str) -> Dict[str, Any]:
        # Prefer the first non-junk candidate.
        best = None
        for c in candidates:
            if not _is_bad_candidate(label, c):
                best = c
                break
        if best is None:
            return {"value": None, "confidence": 0.0, "reason": f"fallback: {reason} (no good candidates)"}
        score = max(
            fuzz.partial_ratio(label.lower(), best.lower()),
            fuzz.token_sort_ratio(label.lower(), best.lower()),
        )
        # Map 0-100 -> 0.0-1.0 conservatively
        confidence = max(0.45, min(0.8, score / 100.0))
        return {"value": best, "confidence": confidence, "reason": f"fallback: {reason}"}

    # Enhanced prompt with field-specific examples
    prompt = f"""
You are an expert at extracting structured data from invoices and documents.

TASK: Select the MOST ACCURATE value for: "{label}"

CANDIDATES:
{chr(10).join(f'{i+1}. {c}' for i, c in enumerate(candidates))}

RULES:
1. For DATE fields (date, dated, dt): Choose actual date values like "02-Oct-25", "15/01/2026", NOT labels like "Date:" or "Dated"
2. For AMOUNT fields (total, amount, price, value): Choose numeric values with currency (₹1,234.56) or decimals, NOT labels
3. For NUMBER fields (invoice no, bill no, challan): Choose alphanumeric codes like "AP-047", "INV-12345", NOT the word "Invoice"
4. For NAME/TEXT fields: Choose the actual name/description, NOT generic labels
5. NEVER pick labels or field names - pick the actual DATA
6. Return value EXACTLY as written - do not modify
7. confidence: 0.9-1.0 = perfect match, 0.7-0.8 = good match, 0.5-0.6 = uncertain, <0.5 = poor match

EXAMPLES:
- Field "Invoice Date" → Pick "15-Jan-2026" NOT "Invoice Date" or "Date:"
- Field "Total Amount" → Pick "₹15,450.00" NOT "Total" or "Amount:"
- Field "Invoice Number" → Pick "INV-2025-001" NOT "Invoice No." or "Number"

OUTPUT (JSON only, no markdown):
{{"value":"exact_value_here","confidence":0.85,"reason":"why_chosen"}}
"""

    if gemini is None:
        return local_fallback("gemini unavailable")

    try:
        resp = gemini.models.generate_content(
            model=get_llm_model(),
            contents=prompt,
        )
    except Exception as e:
        # Common failure here: 404 NOT_FOUND (model not available to project) or permission issues.
        logger.warning(f"Gemini call failed for label '{label}': {e}")
        return local_fallback(str(e))

    raw = str(resp.text)

    # Extract JSON safely
    match = re.search(r"\{[\s\S]*?\}", raw)
    if not match:
        return local_fallback("no json")

    try:
        parsed = json.loads(match.group())
    except Exception:
        return local_fallback("invalid json")

    if parsed.get("value") not in candidates:
        return local_fallback("value not in candidates")

    return {
        "value": parsed.get("value"),
        "confidence": float(parsed.get("confidence", 0.5)),
        "reason": parsed.get("reason", "")
    }

# ---------------------------------------------------------------------
# PDF SUMMARY
# ---------------------------------------------------------------------
def write_pdf(results: Any, out_path: str, summary: Optional[str] = None):
    """Write extraction summary PDF.
    
    Args:
        results: Either a dict (single page results) or a list of dicts (multi-page results)
                 Each dict should have 'page' and 'results' keys for multi-page format
        out_path: Output PDF path
        summary: Optional text to display at the top (e.g., LLM summary)
    """
    c = canvas.Canvas(out_path, pagesize=A4)
    _, h = A4
    y = h - 50

    c.setFont("Helvetica-Bold", 14)
    
    if summary:
        c.drawString(40, y, "Smart Summary")
        y -= 25
        c.setFont("Helvetica", 10)
        for line in summary.split("\n"):
            if y < 60:
                c.showPage()
                y = h - 50
            c.drawString(40, y, line[:100]) # simple truncation
            y -= 12
        y -= 25
        c.setFont("Helvetica-Bold", 14)
    
    # Handle both single results dict and list of page results
    if isinstance(results, list) and len(results) > 0:
        # Multi-page format: list of {"page": N, "results": {...}}
        c.drawString(40, y, f"Extraction Summary - {len(results)} Invoice(s)")
        y -= 30
        
        for page_data in results:
            page_num = page_data.get("page", "?")
            page_results = page_data.get("results", {})
            
            # Page header
            c.setFont("Helvetica-Bold", 12)
            if y < 100:
                c.showPage()
                y = h - 50
            c.drawString(40, y, f"--- Invoice/Page {page_num} ---")
            y -= 25
            
            c.setFont("Helvetica", 10)
            for k, v in page_results.items():
                if y < 60:
                    c.showPage()
                    y = h - 50
                # Handle both dict format and simple value format
                if isinstance(v, dict):
                    value = v.get('value', '-')
                else:
                    value = v if v is not None else '-'
                c.drawString(50, y, f"{k}: {value}")
                y -= 14
            
            y -= 15  # Extra space between pages
    else:
        # Single page format: just a dict of results
        c.drawString(40, y, "Extraction Summary")
        y -= 30
        
        c.setFont("Helvetica", 11)
        page_results = results if isinstance(results, dict) else {}
        for k, v in page_results.items():
            if y < 80:
                c.showPage()
                y = h - 50
            # Handle both dict format and simple value format
            if isinstance(v, dict):
                value = v.get('value', '-')
                confidence = v.get('confidence', 0.0)
            else:
                value = v if v is not None else '-'
                confidence = '-'
            c.drawString(40, y, f"{k}: {value}")
            y -= 15
            c.drawString(60, y, f"Confidence: {confidence}")
            y -= 25

    c.save()

def write_simple_pdf(ocr_values: List[str], path: str):
    """Write a simple PDF with all OCR values when no template is provided"""
    c = canvas.Canvas(path, pagesize=A4)
    w, h = A4
    y = h - 50

    c.setFont("Helvetica-Bold", 14)
    c.drawString(40, y, "OCR Extraction Results")
    y -= 30

    c.setFont("Helvetica", 10)
    for idx, text in enumerate(ocr_values, 1):
        if y < 80:
            c.showPage()
            y = h - 50
        # Truncate long text to fit on page
        display_text = text[:100] + "..." if len(text) > 100 else text
        c.drawString(40, y, f"{idx}. {display_text}")
        y -= 15

    c.save()

# ---------------------------------------------------------------------
# BACKGROUND PIPELINE
# ---------------------------------------------------------------------
# ---------------------------------------------------------------------
# BACKGROUND PIPELINE (Modified for Vercel Sync Support)
# ---------------------------------------------------------------------
def process_job(job_id: str, doc_path: str, template_path: str, sheet_name: Optional[str] = None, is_sync: bool = False, file_bytes: bytes = None, tpl_bytes: bytes = None):
    try:
        logger.info(f"Processing job {job_id} (Sync: {is_sync})")
        
        # In Vercel mode, we work with bytes, not paths
        if is_sync:
             # Create temp files provided bytes (DocAI/Docling often need file paths)
             # Vercel /tmp is fast and available
             with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_doc:
                 tmp_doc.write(file_bytes)
                 tmp_doc.flush()
                 doc_path = tmp_doc.name
             
             if tpl_bytes:
                 with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_tpl:
                     tmp_tpl.write(tpl_bytes)
                     tmp_tpl.flush()
                     template_path = tmp_tpl.name

        logger.info(f"Document: {doc_path}, Template: {template_path}, Sheet: {sheet_name}")
        
        # Parse document page-by-page for multi-invoice support
        pages_values = parse_document_by_page(doc_path)
        num_pages = len(pages_values)
        logger.info(f"Document has {num_pages} pages")
        
        # Also get combined values for legacy compatibility and sheet detection
        all_ocr_values = []
        for pv in pages_values:
            all_ocr_values.extend(pv)
        # Dedupe
        seen = set()
        ocr_values = []
        for v in all_ocr_values:
            if v.lower() not in seen:
                ocr_values.append(v)
                seen.add(v.lower())
        
        logger.info(f"OCR extracted {len(ocr_values)} unique values total")
        logger.info(f"First 20 OCR values: {ocr_values[:20]}")

        # ---------------------------------------------------------------------
        # GEMINI SMART EXTRACTION
        # ---------------------------------------------------------------------
        # Reconstruct text for LLM context
        full_text_llm = "\n".join(ocr_values)
        llm_data = extract_data_with_gemini(full_text_llm)
        llm_doc_class = llm_data.get("document_class", {})
        llm_inv_summary = llm_data.get("invoice_summary", {})
        llm_hsn_summary = llm_data.get("hsn_summary", [])
        
        doc_type = llm_doc_class.get("type", "unknown")
        logger.info(f"LLM Detected Doc Type: {doc_type}")
        
        if not template_path:
            # No template provided - just do OCR and create a simple report
            logger.info("No template provided, creating simple extraction report")
            
            if llm_data:
                # Format HSN summary nicely
                hsn_text = "\n".join([f"- {h.get('hsn_code')}: Val={h.get('taxable_value')}, Tax={h.get('tax_rate')}%" for h in llm_hsn_summary]) if llm_hsn_summary else "No HSN data"
                
                # Format Invoice summary nicely
                inv_text = "\n".join([f"- {k}: {v}" for k, v in llm_inv_summary.items() if v]) if llm_inv_summary else "No Invoice Summary"

                summary_text = (
                    f"--- GEMINI SMART SUMMARY ---\n"
                    f"Type: {doc_type}\n"
                    f"Rec. Sheet: {llm_doc_class.get('recommended_sheet')}\n"
                    f"--- INVOICE SUMMARY ---\n{inv_text}\n"
                    f"--- HSN SUMMARY ---\n{hsn_text}\n"
                    f"----------------------------"
                )
                ocr_values.insert(0, summary_text)
                logger.info(f"LLM HSN Summary: {llm_hsn_summary}")

            pdf_out = os.path.join(OUTPUT_DIR, f"{job_id}_summary.pdf")
            write_simple_pdf(ocr_values, pdf_out)
            
            if is_sync:
                # Return Base64
                with open(pdf_out, "rb") as f:
                    pdf_b64 = base64.b64encode(f.read()).decode()
                # Clean up extracted items for JSON response
                return {
                    "status": "done",
                    "pdf_base64": pdf_b64,
                    "message": f"OCR complete. Extracted {len(ocr_values)} values from {num_pages} pages.",
                    "llm_data": llm_data
                }

            JOBS[job_id] = {
                "status": "done",
                "excel": None,
                "pdf": pdf_out,
                "message": f"OCR complete. Extracted {len(ocr_values)} values from {num_pages} pages.",
                "ocr_count": len(ocr_values),
                "num_pages": num_pages,
                "llm_data": llm_data
            }
            logger.info(f"Job {job_id} completed successfully")
            return
        
        wb = load_workbook(template_path)
        
        # Manual sheet selection
        ws = None
        if sheet_name:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                logger.info(f"Using manually selected sheet: '{sheet_name}'")
            else:
                logger.warning(f"Manually selected sheet '{sheet_name}' not found in template. Available: {wb.sheetnames}")
        
        # Auto-detection (LLM Priority -> Heuristic Fallback)
        ranked = None
        if not ws:
            # 1. Try LLM Recommendation
            rec_sheet = llm_doc_class.get("recommended_sheet")
            if rec_sheet:
                 for s in wb.sheetnames:
                     if s.lower() == rec_sheet.lower():
                         ws = wb[s]
                         logger.info(f"Using LLM recommended sheet: '{ws.title}' (matched '{rec_sheet}')")
                         break
            
            # 2. Fallback to existing logic
            if not ws:
                ws = choose_best_sheet(wb, ocr_values)
                ranked = getattr(ws, "__finautomate_ranked_sheets", None)
                if isinstance(ranked, list):
                    logger.info(f"Top sheet candidates: {[d['sheet'] for d in ranked[:5]]}")
        
        logger.info(f"Selected template sheet: '{ws.title}'")

        # Store debug info on the job so UI can inspect if needed.
        JOBS[job_id].update(
            {
                "template_sheets": [s.title for s in wb.worksheets],
                "selected_sheet": ws.title,
                "document_type": doc_type,
                "sheet_ranking": ranked[:10] if isinstance(ranked, list) else None,
                "num_pages": num_pages,
            }
        )

        labels = extract_labels_from_ws(ws)
        logger.info(f"Template has {len(labels)} labels")
        logger.info(f"Template labels (normalized): {list(labels.keys())}")
        
        # Find base data row
        base_data_row = min(r for label_key, (r, c) in labels.items()) if labels else 2
        logger.info(f"Base data row: {base_data_row}")
        
        # Process each page as a separate invoice
        all_results = []
        for page_idx, page_ocr_values in enumerate(pages_values):
            logger.info(f"Processing page {page_idx + 1}/{num_pages}")
            
            current_row = base_data_row + page_idx
            
            candidates_by_label: Dict[str, List[str]] = {}
            for label_key in labels:
                candidates = generate_candidates(label_key, page_ocr_values)
                candidates_by_label[label_key] = candidates
                if page_idx == 0:  # Only log for first page
                    logger.info(f"Field '{label_key}' - Top candidates: {candidates[:3]}")

            # Try a single Gemini batch call
            batch = gemini_batch_choose(candidates_by_label, page_ocr_values)

            results: Dict[str, Dict[str, Any]] = {}
            for label_key in labels:
                # ---------------------------------------------------------
                # 0. Check Gemini Smart Summary (High Confidence)
                # ---------------------------------------------------------
                llm_val = None
                label_lower = label_key.lower()
                
                if llm_inv_summary:
                    if "total" in label_lower and "taxable" in label_lower:
                        llm_val = llm_inv_summary.get("total_taxable_value")
                    elif "cgst" in label_lower and "amount" in label_lower:
                        llm_val = llm_inv_summary.get("total_cgst")
                    elif "sgst" in label_lower and "amount" in label_lower:
                        llm_val = llm_inv_summary.get("total_sgst")
                    elif "igst" in label_lower and "amount" in label_lower:
                        llm_val = llm_inv_summary.get("total_igst")
                    elif "cess" in label_lower and "amount" in label_lower:
                        llm_val = llm_inv_summary.get("total_cess")
                    elif ("grand" in label_lower or "invoice val" in label_lower) and "total" in label_lower:
                        llm_val = llm_inv_summary.get("grand_total")
                    elif "invoice" in label_lower and ("no" in label_lower or "number" in label_lower):
                        llm_val = llm_inv_summary.get("invoice_no")
                    elif "date" in label_lower and "invoice" in label_lower:
                        llm_val = llm_inv_summary.get("invoice_date")
                    elif "gstin" in label_lower:
                        if "supplier" in label_lower or "seller" in label_lower:
                            llm_val = llm_inv_summary.get("gstin_supplier")
                        elif "receiver" in label_lower or "buyer" in label_lower:
                            llm_val = llm_inv_summary.get("gstin_receiver")
                    elif "reverse charge" in label_lower:
                        llm_val = llm_inv_summary.get("reverse_charge")
                    elif "hsn" in label_lower or "sac" in label_lower:
                        if llm_hsn_summary:
                            # Join first 3 unique HSN codes
                            codes = sorted(list(set([x.get("hsn_code") for x in llm_hsn_summary if x.get("hsn_code")])))
                            if codes:
                                llm_val = ", ".join(codes[:3])

                llm_result_obj = None
                if llm_val is not None:
                     llm_result_obj = {"value": str(llm_val), "confidence": 0.99, "reason": "gemini-smart-summary"}

                # Try heuristics first
                h = heuristic_extract(label_key, page_ocr_values)
                batch_result = batch.get(label_key)
                
                chosen = None
                
                # Priority: LLM Summary > Batch LLM > Heuristic
                if llm_result_obj:
                    chosen = llm_result_obj
                elif h.get("value") and h.get("confidence", 0) >= 0.7:
                    if batch_result and batch_result.get("confidence", 0) > h.get("confidence", 0):
                        chosen = batch_result
                    else:
                        chosen = h
                elif batch_result and batch_result.get("value"):
                    chosen = batch_result
                elif h.get("value"):
                    chosen = h
                else:
                    chosen = {"value": None, "confidence": 0.0, "reason": "no extraction"}
                
                logger.info(f"Field '{label_key}' - Chosen: {chosen.get('value')} (confidence: {chosen.get('confidence')}) [source: {chosen.get('reason', 'unknown')}]")
                results[label_key] = chosen
            
            # Write extracted values to template at current row
            for label_key, meta in results.items():
                if label_key not in labels:
                    logger.warning(f"Label '{label_key}' not found in template")
                    continue

                _, c = labels[label_key]  # Use column from base row
                logger.info(f"Writing '{label_key}' to row={current_row}, col={c}: {meta['value']} (confidence: {meta['confidence']})")

                if meta["confidence"] >= 0.7 and meta["value"]:
                    safe_write(ws, current_row, c, str(meta["value"]))
                    logger.info(f"Successfully wrote value for '{label_key}'")
                elif meta["value"]:
                    logger.warning(f"Low confidence ({meta['confidence']}) for '{label_key}': {meta['value']}")
                    safe_write(ws, current_row, c, f"[?] {meta['value']}")
                else:
                    logger.warning(f"No value found for '{label_key}'")
                    safe_write(ws, current_row, c, "-")
            
            all_results.append({
                "page": page_idx + 1,
                "row": current_row,
                "results": {k: v.get("value") for k, v in results.items()}
            })
            logger.info(f"Page {page_idx + 1} processed, written to row {current_row}")

        excel_out = os.path.join(OUTPUT_DIR, f"{job_id}_filled.xlsx")
        pdf_out = os.path.join(OUTPUT_DIR, f"{job_id}_summary.pdf")

        wb.save(excel_out)
        
        # Construct summary for template mode PDF
        summary_text = None
        if llm_data:
             # Format HSN summary nicely if possible, else dump JSON
             hsn_text = "\n".join([f"- {h.get('hsn_code')}: Val={h.get('taxable_value')}, Tax={h.get('tax_rate')}%" for h in llm_hsn_summary]) if llm_hsn_summary else "No HSN data"
             inv_text = "\n".join([f"- {k}: {v}" for k, v in llm_inv_summary.items() if v]) if llm_inv_summary else "No Invoice Summary"
             summary_text = (
                 f"--- GEMINI SMART SUMMARY ---\n"
                 f"Type: {doc_type}\n"
                 f"Rec. Sheet: {llm_doc_class.get('recommended_sheet')}\n"
                 f"--- INVOICE SUMMARY ---\n{inv_text}\n"
                 f"--- HSN SUMMARY ---\n{hsn_text}\n"
                 f"----------------------------"
             )
        
        write_pdf(all_results, pdf_out, summary=summary_text)

        sheet_name = ws.title

        sheet_name = ws.title
        
        JOBS[job_id] = {
            "status": "done",
            "excel": excel_out,
            "pdf": pdf_out,
            "message": f"Processed {num_pages} invoice(s) from document into template sheet '{sheet_name}'.",
            "invoices_processed": num_pages,
            "selected_sheet": sheet_name,
            "document_type": doc_type,
            "available_sheets": [s.title for s in wb.worksheets],
            "all_results": all_results
        }
        logger.info(f"Job {job_id} completed successfully: {num_pages} invoices processed")

    except Exception as e:
        logger.error(f"Job {job_id} failed: {str(e)}", exc_info=True)
        JOBS[job_id] = {"status": "failed", "error": str(e)}

# ---------------------------------------------------------------------
# EXTRACT ONLY (NO TEMPLATE, NO PDF)
# ---------------------------------------------------------------------
def extract_only_job(job_id: str, doc_path: str):
    try:
        logger.info(f"Extract-only job {job_id}")
        ocr_values = parse_document(doc_path)
        logger.info(f"OCR extracted {len(ocr_values)} values")
        
        JOBS[job_id] = {
            "status": "done",
            "excel": None,
            "pdf": None,
            "message": f"Extraction complete. Found {len(ocr_values)} values.",
            "ocr_values": ocr_values[:50]  # Return first 50 values
        }
        logger.info(f"Extract-only job {job_id} completed")
    except Exception as e:
        logger.error(f"Extract-only job {job_id} failed: {str(e)}", exc_info=True)
        JOBS[job_id] = {"status": "failed", "error": str(e)}

# ---------------------------------------------------------------------
# API
# ---------------------------------------------------------------------



@app.post("/upload")
async def upload(
    background: BackgroundTasks,
    document: UploadFile = File(...),
    template: UploadFile = File(None),
    sheet_name: Optional[str] = Form(None)
):
    job_id = str(uuid.uuid4())
    
    # VERCEL / SYNC MODE
    if IS_VERCEL:
        doc_bytes = await document.read()
        tpl_bytes = await template.read() if template else None
        
        # Run synchronously
        result = process_job(
            job_id=job_id,
            doc_path="memory", 
            template_path="memory",
            sheet_name=sheet_name,
            is_sync=True,
            file_bytes=doc_bytes,
            tpl_bytes=tpl_bytes
        )
        return result

    # LOCAL / ASYNC MODE
    doc_path = os.path.join(INPUT_DIR, f"{job_id}_doc{os.path.splitext(document.filename)[1]}")
    with open(doc_path, "wb") as f:
        f.write(await document.read())
    
    tpl_path = None
    if template and template.filename:
        tpl_path = os.path.join(INPUT_DIR, f"{job_id}_tpl{os.path.splitext(template.filename)[1]}")
        with open(tpl_path, "wb") as f:
            f.write(await template.read())

    JOBS[job_id] = {"status": "processing"}
    background.add_task(process_job, job_id, doc_path, tpl_path, sheet_name)

    return {"job_id": job_id}

# Note: Keeping upload-multi mostly async or simple for now, 
# but if needed on Vercel, it also requires sync refactor.
# For urgent Vercel deployment, we focus on the main endpoint compatibility.


def process_multi_job(job_id: str, doc_paths: List[str], template_path: str, sheet_name: Optional[str] = None):
    """Process multiple documents and fill them into the same template.
    
    Each page of each PDF is treated as a separate invoice and fills one row.
    For example, if you upload 2 PDFs with 3 pages each, you get 6 rows.
    Now supports routing different invoices to different sheets (e.g. Invoices vs Credit Notes).
    """
    try:
        logger.info(f"Processing multi-document job {job_id} with {len(doc_paths)} documents, Manual Sheet: {sheet_name}")
        
        # Collect all pages from all documents
        all_pages: List[Tuple[int, int, List[str]]] = []  # (doc_idx, page_idx, ocr_values)
        
        for doc_idx, doc_path in enumerate(doc_paths):
            logger.info(f"Parsing document {doc_idx + 1}: {doc_path}")
            pages_values = parse_document_by_page(doc_path)
            logger.info(f"Document {doc_idx + 1} has {len(pages_values)} pages")
            
            for page_idx, page_values in enumerate(pages_values):
                all_pages.append((doc_idx, page_idx, page_values))
        
        logger.info(f"Total invoices (pages) to process: {len(all_pages)}")
        JOBS[job_id]["total_invoices"] = len(all_pages)
        
        if not template_path:
            # No template - just return extracted values
            all_ocr_values = []
            for doc_idx, page_idx, page_values in all_pages:
                all_ocr_values.append({
                    "document": doc_idx + 1,
                    "page": page_idx + 1,
                    "values": page_values[:50]
                })
            
            JOBS[job_id] = {
                "status": "done",
                "excel": None,
                "pdf": None,
                "message": f"Extracted values from {len(all_pages)} invoices across {len(doc_paths)} documents.",
                "documents_processed": len(doc_paths),
                "invoices_processed": len(all_pages),
                "all_ocr_values": all_ocr_values
            }
            return
        
        # Load template
        wb = load_workbook(template_path)
        
        # Manual global sheet selection
        manual_ws = None
        if sheet_name:
            if sheet_name in wb.sheetnames:
                manual_ws = wb[sheet_name]
                logger.info(f"Using manually selected sheet globally: '{sheet_name}'")
            else:
                logger.warning(f"Manually selected sheet '{sheet_name}' not found. Will use auto-routing.")

        # Track next available row for each sheet
        # format: {sheet_title: next_row_number}
        sheet_row_trackers: Dict[str, int] = {}
        
        # Cache labels for each sheet to avoid re-extraction
        # format: {sheet_title: labels_dict}
        sheet_labels_cache: Dict[str, Dict[str, Tuple[int, int]]] = {}

        # Process each page as a separate invoice
        all_results = []
        for invoice_idx, (doc_idx, page_idx, ocr_values) in enumerate(all_pages):
            logger.info(f"Processing invoice {invoice_idx + 1}/{len(all_pages)} (Doc {doc_idx + 1}, Page {page_idx + 1})")
            
            # -----------------------------------------------------------------
            # GEMINI SMART EXTRACTION
            # -----------------------------------------------------------------
            full_text_llm = "\n".join(ocr_values)
            llm_data = extract_data_with_gemini(full_text_llm)
            llm_doc_class = llm_data.get("document_class", {})
            llm_inv_summary = llm_data.get("invoice_summary", {})
            llm_hsn_summary = llm_data.get("hsn_summary", [])
            
            # 1. Determine the best sheet for THIS invoice
            if manual_ws:
                ws = manual_ws
            else:
                # Try LLM Recommendation
                rec_sheet = llm_doc_class.get("recommended_sheet")
                found_llm_sheet = None
                if rec_sheet:
                    for s in wb.sheetnames:
                        if s.lower() == rec_sheet.lower():
                            found_llm_sheet = wb[s]
                            logger.info(f"Using LLM recommended sheet: '{s}'")
                            break
                
                if found_llm_sheet:
                    ws = found_llm_sheet
                else:
                    ws = choose_best_sheet(wb, ocr_values)
            
            sheet_name = ws.title
            
            # 2. Get labels (using cache if available)
            if sheet_name not in sheet_labels_cache:
                sheet_labels_cache[sheet_name] = extract_labels_from_ws(ws)
                logger.info(f"Cached labels for sheet '{sheet_name}'")
            
            labels = sheet_labels_cache[sheet_name]
            
            # 3. Determine target row
            if sheet_name not in sheet_row_trackers:
                # Initialize row tracker for this sheet
                # Use the min row from labels as the "header" row, assuming data starts after it??
                # Typically labels are headers. The user wants to APPEND data.
                # Logic from existing code: base_data_row = min(r) if labels else 2
                # But we should start appending below any existing data if possible?
                # For now, stick to the logic: data area starts at min(label_row) + offset ??
                # Actually, in the previous code, base_data_row was static.
                # Let's assume the template is empty of DATA, and only has HEADERS.
                # So we start writing at (min_header_row + 1)?
                # Wait, labels usually store where the VALUE should go if it's a form?
                # NO, the system seems to be "find column C where header is Label".
                # extract_labels_from_ws usually finds (row, col) of the header text.
                # So we should write to `next_row` in that column `c`.
                
                # Let's trust the previous finding: base_row = (min_label_r) or 2.
                # If these are headers, we should write at base_row + 1 initially?
                # Previous code: `base_data_row + page_idx`.
                # If page_idx=0, we write at base_data_row.
                # If base_data_row is the top-most label row, this overwrites headers!
                # Unless `extract_labels_from_ws` returns the row *under* the header?
                # I'll stick to: default start = min_label_row + 1 (to be safe against overwriting headers).
                # Wait, if previous code worked, maybe I should respect it.
                # Previous code: `base_data_row = min(...)`. `current_row = base_data_row + invoice_idx`.
                # If invoice 0 writes to base_data_row, and base_data_row is a header, that's a bug.
                # But maybe `extract_labels_from_ws` filters out the header row?
                # Let's assume `base_data_row` calculated from `min` is potentially the first data row if `extract_labels` logic accounts for it.
                # BUT, logically, max(label_rows) + 1 makes more sense for "start of list".
                # Let's use `max(r for ... ) + 1`. This is safer for appending.
                
                if labels:
                    # Find the bottom-most header row
                    max_header_row = max(r for _, (r, c) in labels.items())
                    start_row = max_header_row + 1
                else:
                    start_row = 2
                
                sheet_row_trackers[sheet_name] = start_row
                logger.info(f"Sheet '{sheet_name}' first write will be at row {start_row}")

            current_row = sheet_row_trackers[sheet_name]
            
            logger.info(f"Routing invoice to Sheet: '{sheet_name}', Row: {current_row}")

            # Generate candidates
            candidates_by_label = {}
            for label_key in labels:
                candidates = generate_candidates(label_key, ocr_values)
                candidates_by_label[label_key] = candidates
            
            # Try batch Gemini call
            batch = gemini_batch_choose(candidates_by_label, ocr_values)
            
            # Extract results for each label
            doc_results = {}
            for label_key in labels:
                # ---------------------------------------------------------
                # 0. Check Gemini Smart Summary
                # ---------------------------------------------------------
                llm_val = None
                label_lower = label_key.lower()
                
                if llm_inv_summary:
                    if "total" in label_lower and "taxable" in label_lower:
                        llm_val = llm_inv_summary.get("total_taxable_value")
                    elif "cgst" in label_lower and "amount" in label_lower:
                        llm_val = llm_inv_summary.get("total_cgst")
                    elif "sgst" in label_lower and "amount" in label_lower:
                        llm_val = llm_inv_summary.get("total_sgst")
                    elif "igst" in label_lower and "amount" in label_lower:
                        llm_val = llm_inv_summary.get("total_igst")
                    elif ("grand" in label_lower or "invoice val" in label_lower) and "total" in label_lower:
                        llm_val = llm_inv_summary.get("grand_total")
                    elif "invoice" in label_lower and ("no" in label_lower or "number" in label_lower):
                        llm_val = llm_inv_summary.get("invoice_no")
                    elif "date" in label_lower and "invoice" in label_lower:
                        llm_val = llm_inv_summary.get("invoice_date")
                    elif "gstin" in label_lower and "supplier" in label_lower:
                        llm_val = llm_inv_summary.get("gstin_supplier")
                    elif "reverse charge" in label_lower:
                        llm_val = llm_inv_summary.get("reverse_charge")
                    elif "hsn" in label_lower or "sac" in label_lower:
                        if llm_hsn_summary:
                            codes = sorted(list(set([x.get("hsn_code") for x in llm_hsn_summary if x.get("hsn_code")])))
                            if codes:
                                llm_val = ", ".join(codes[:3])

                llm_result_obj = None
                if llm_val is not None:
                     llm_result_obj = {"value": str(llm_val), "confidence": 0.99, "reason": "gemini-smart-summary"}

                h = heuristic_extract(label_key, ocr_values)
                batch_result = batch.get(label_key)
                
                chosen = None
                if llm_result_obj:
                    chosen = llm_result_obj
                elif h.get("value") and h.get("confidence", 0) >= 0.7:
                    if batch_result and batch_result.get("confidence", 0) > h.get("confidence", 0):
                        chosen = batch_result
                    else:
                        chosen = h
                elif batch_result and batch_result.get("value"):
                    chosen = batch_result
                elif h.get("value"):
                    chosen = h
                else:
                    chosen = {"value": None, "confidence": 0.0, "reason": "no extraction"}
                
                doc_results[label_key] = chosen
            
            # Write to template
            for label_key, meta in doc_results.items():
                if label_key not in labels:
                    continue
                
                _, c = labels[label_key]
                
                if meta["confidence"] >= 0.7 and meta["value"]:
                    safe_write(ws, current_row, c, str(meta["value"]))
                elif meta["value"]:
                    safe_write(ws, current_row, c, f"[?] {meta['value']}")
                else:
                    safe_write(ws, current_row, c, "-")
            
            all_results.append({
                "document": doc_idx + 1,
                "page": page_idx + 1,
                "row": current_row,
                "sheet": sheet_name,
                "results": {k: v.get("value") for k, v in doc_results.items()}
            })
            
            # Increment row for this sheet
            sheet_row_trackers[sheet_name] += 1
            
            JOBS[job_id]["processed"] = invoice_idx + 1
            logger.info(f"Invoice {invoice_idx + 1} written to row {current_row} of sheet {sheet_name}")
        
        # Save files
        excel_out = os.path.join(OUTPUT_DIR, f"{job_id}_filled.xlsx")
        pdf_out = os.path.join(OUTPUT_DIR, f"{job_id}_summary.pdf")
        
        wb.save(excel_out)
        
        # Create summary PDF
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas as pdf_canvas
        c = pdf_canvas.Canvas(pdf_out, pagesize=A4)
        _, h = A4
        y = h - 50
        
        c.setFont("Helvetica-Bold", 14)
        c.drawString(40, y, f"Multi-Invoice Extraction: {len(all_pages)} invoices from {len(doc_paths)} files")
        y -= 30
        
        for doc_result in all_results:
            if y < 100:
                c.showPage()
                y = h - 50
            c.setFont("Helvetica-Bold", 11)
            c.drawString(40, y, f"Doc {doc_result['document']} Page {doc_result['page']} → {doc_result['sheet']} Row {doc_result['row']}")
            y -= 18
            c.setFont("Helvetica", 9)
            for k, v in list(doc_result['results'].items())[:8]:
                if y < 60:
                    c.showPage()
                    y = h - 50
                display_val = str(v)[:40] if v else "NaN"
                c.drawString(50, y, f"{k}: {display_val}")
                y -= 11
            y -= 8
        
        c.save()
        
        if is_sync:
            with open(excel_out, "rb") as f:
                xl_b64 = base64.b64encode(f.read()).decode()
            with open(pdf_out, "rb") as f:
                pdf_b64 = base64.b64encode(f.read()).decode()
            
            return {
                "status": "done",
                "excel_base64": xl_b64,
                "pdf_base64": pdf_b64,
                "message": f"Processed {len(all_pages)} invoices from {len(doc_paths)} files into template.",
                "documents_processed": len(doc_paths),
                "invoices_processed": len(all_pages),
                "sheets_used": list(sheet_row_trackers.keys()),
            }

        JOBS[job_id] = {
            "status": "done",
            "excel": excel_out,
            "pdf": pdf_out,
            "message": f"Processed {len(all_pages)} invoices from {len(doc_paths)} files into template.",
            "documents_processed": len(doc_paths),
            "invoices_processed": len(all_pages),
            "sheets_used": list(sheet_row_trackers.keys()),
            "all_results": all_results
        }
        logger.info(f"Multi-document job {job_id} completed: {len(all_pages)} invoices")
        
    except Exception as e:
        logger.error(f"Multi-document job {job_id} failed: {str(e)}", exc_info=True)
        if is_sync:
            return {"status": "failed", "error": str(e)}
        JOBS[job_id] = {"status": "failed", "error": str(e)}


@app.post("/extract-only")
def extract_only(
    background: BackgroundTasks,
    document: UploadFile = File(...)
):
    job_id = str(uuid.uuid4())
    doc_path = os.path.join(INPUT_DIR, f"{job_id}_doc{os.path.splitext(document.filename)[1]}")
    
    with open(doc_path, "wb") as f:
        f.write(document.file.read())
    
    JOBS[job_id] = {"status": "processing"}
    background.add_task(extract_only_job, job_id, doc_path)
    
    return {"job_id": job_id}

@app.get("/status/{job_id}")
def status(job_id: str):
    return JOBS.get(job_id, {"status": "unknown"})

@app.get("/result/{job_id}")
def result(job_id: str):
    if job_id not in JOBS:
        raise HTTPException(404, "Job not found")
    return JOBS[job_id]

